import dash
from datetime import datetime
from dash import html, dcc, Input, Output
from dash.dependencies import Input, Output, State
from openpyxl import load_workbook
from datetime import date
import pandas as pd
import dash_bootstrap_components as dbc
import plotly.graph_objects as go
import dash_ag_grid as dag
from datetime import date
from dash import html
from layouthome_2 import *
from layoutlogin import *
from layoutagende import *
from tempidiattesa import *
from commonlayout import *
from consiglidimensionamento import *
from ulterioridettagli_2 import *
from generazione_analisi_scheduler_per_frontend import *
from roots import *
#import os
import io
#import base64  
import time



dtype={'DIARY_ID': str,
       'SLOTS_MASK': str,
       'TIMEBAND_ID': str
       }

dtype_2 = {'CODICE_AGENDA': str,
    'STS11': str,
    'CODICE_EROGANTE': str,
    }

start_time = time.time()
###############################################
## BASI DATI NEL CASO SI DEVONO AGGIORNARE #####
###############################################
raggruppam = pd.read_csv(os.path.join(base_path, 'File_utili', 'calendar_7_agende_selezionate_con_raggruppamento.csv'), sep =',')
cluster = pd.read_excel(os.path.join(base_path, 'File_utili', 'Clustering_aderenza_grafo_new_600.xlsx'))
cluster.rename(columns ={'Prestazione' : 'Codice Prestazione'},inplace=True)
#calendar_agende = pd.read_csv(os.path.join(base_path, 'File_utili', 'V_AGENDE_ATTIVE_CE_CALENDAR.csv'), sep=";", dtype=dtype)
calendar_agende = pd.read_parquet(os.path.join(base_path, 'File_utili', 'V_AGENDE_ATTIVE_CE_CALENDAR.parquet'))
## TABELLA PRINCIPALE 
#priorita = pd.read_csv(os.path.join(base_path, 'File_utili', 'V_AGENDE_ATTIVE_CE.csv'), sep=';')
priorita = pd.read_parquet(os.path.join(base_path, 'File_utili', 'V_AGENDE_ATTIVE_CE.parquet'))


prenot = pd.read_parquet(os.path.join(base_path, 'File_utili', 'cup_prenotato_cleaned_wave_1.parquet'))
re_cluster = pd.read_excel(os.path.join(base_path, 'File_utili', 'Re-clustering_Spec_Fa.Re_v.2.0_20240624.xlsx'), sheet_name = 'Spectral_clustering_600_limited')
decodifica_catalogo = pd.read_csv(os.path.join(base_path,'File_utili','decodifica_descrizioni_catalogo.csv'),sep=';')
sentinella =pd.read_excel(os.path.join(base_path,'File_utili','Prestazioni sentinella_v1.0.xlsx'), sheet_name='Foglio 1')
edge_list_completa_22_23 = pd.read_parquet(os.path.join(base_path, 'File_utili', 'edge_list_completa_22_23.parquet'))
####prestazioni_agende_df = pd.read_csv(os.path.join(base_path, 'Df_demo','prestazioni_agende.csv'), sep =',')
decodifica_ambiti_garanzia = pd.read_csv(os.path.join(base_path, 'File_utili', 'decodifica_ambiti_garanzia 2.csv'),sep=';')
#timeband_orders_df = pd.read_csv(os.path.join(base_path,'File_utili','V_AGENDE_ATTIVE_CE_TIMEBAND_ORDERS.csv'), sep =';' , encoding='ISO-8859-1', dtype=dtype_2) #"ISO-8859-1",        , dtype=dtype
timeband_orders_df = pd.read_parquet(os.path.join(base_path,'File_utili','V_AGENDE_ATTIVE_CE_TIMEBAND_ORDERS.parquet'))
timeband_orders_df = timeband_orders_df[~(timeband_orders_df['ORDER_ID'].str.endswith('_5', na=False))] 

df_dettaglio_struttura = pd.read_csv(root_path+r"\df_dettaglio_struttura.csv", sep =',')
## DF CALCOLATI

## DF DI TUTTE LE AGENDE PRESE IN CONSIDERAZIONE CON DESCRIZIONE 
agende_descrizione = priorita[['CODICE_AGENDA','DESCRIZIONE_AGENDA']].drop_duplicates()
agende_descrizione.rename(columns = {'CODICE_AGENDA' : 'Codice Agenda','DESCRIZIONE_AGENDA' : 'Descrizione Agenda'},inplace=True)

### decodifica descrizione recluster
decodifica_descrizione = re_cluster[['Re-clustering','Descrizione post-Reclustering']].drop_duplicates()
decodifica_descrizione['Re-clustering'] = decodifica_descrizione['Re-clustering'].astype(str)


### prestazioni_agende_df
df_agende_no_prestazioni_nuove_agende_attive = priorita[~(priorita['CODICE_PREST_SERV'].str.endswith('_5', na=False))] 
df_agende_no_prestazioni_nuove_agende_attive = df_agende_no_prestazioni_nuove_agende_attive[['CODICE_AGENDA','CODICE_PREST_SERV','DESCR_PREST']].drop_duplicates()
prestazioni_agende_df = df_agende_no_prestazioni_nuove_agende_attive.rename(columns = {'CODICE_AGENDA' : 'Codice Agenda', 'CODICE_PREST_SERV' : 'Codice Prestazione' ,'DESCR_PREST' : 'Descrizione Prestazione' })
prestazioni_agende_df['Codice Agenda'] = prestazioni_agende_df['Codice Agenda'].astype(str)

# per funzionamento calcolo ambito dal file excel caricato
priorita_nuovo = priorita.drop_duplicates(subset=['CODICE_AGENDA', 'CODICE_EROGANTE', 'STS11'])
priorita_nuovo = priorita_nuovo[['CODICE_AGENDA', 'CODICE_EROGANTE', 'STS11']]

def import_funzione_fe():

    scheduler_a01 = load_scheduler(file_path = SCHEDULER_ABSOLUTE_PATH+r'\OUTPUT\my_scheduler_ASL Caserta A01_20240627.pkl')
    scheduler_a02 = load_scheduler(file_path = SCHEDULER_ABSOLUTE_PATH+r'\OUTPUT\my_scheduler_ASL Caserta A02_20240627.pkl')
    scheduler_a03 = load_scheduler(file_path = SCHEDULER_ABSOLUTE_PATH+r'\OUTPUT\my_scheduler_ASL Caserta A03_20240627.pkl')
    scheduler_a04 = load_scheduler(file_path = SCHEDULER_ABSOLUTE_PATH+r'\OUTPUT\my_scheduler_ASL Caserta A04_20240627.pkl')

    # CAPIRE SE SERVE DF_PREVISIONI
    #df_previsioni = import_previsioni_cluster(input_folder_location=SCHEDULER_ABSOLUTE_PATH+f'\INPUT\previsioni_domanda\{ambito}\\')
    #df_previsioni = transform_previsioni(df_previsioni, start_day)
    # Import agende con info ambito
    df_agende_da_suddividere = import_csv(SCHEDULER_ABSOLUTE_PATH+r"/INPUT/agende_da_suddividere.csv")

    df_agende_unique_fe = import_agende_ambito(df_agende_da_suddividere, SCHEDULER_ABSOLUTE_PATH+r"/INPUT/")    ## seerve??
    df_cluster = load_clustering_data(SCHEDULER_ABSOLUTE_PATH+r"/INPUT/")
    #dict_duration_cluster = filter_clusters(df_agende_unique, df_cluster, ambito, lambda x: x.mode().values[0])
    calendar_fe = import_calendar(SCHEDULER_ABSOLUTE_PATH+r'\INPUT\\')                                        ## seerve??
    file_c_df = read_file_c(cartella_file_sorgente = SCHEDULER_ABSOLUTE_PATH+r"/INPUT/")                   ## seerve??
    df_reclustering = load_reclustering_data(file_path=SCHEDULER_ABSOLUTE_PATH+r"/INPUT/", file_name="Re-clustering_Spec_Fa.Re_v.2.0_20240624.xlsx", sheet_name='Spectral_clustering_600_limited')
    dict_peso_prestazioni_tutto = ottieni_conteggio_prestazioni_tutto(file_c_df)

    df_reclustering = process_reclustering_data(df=df_reclustering, weight_dict=dict_peso_prestazioni_tutto)
    df_cluster = cluster_add_pesi(df_cluster, dict_peso_prestazioni_tutto)
    df_cup_eapoc = load_cup_eapoc(df_reclustering, input_path=SCHEDULER_ABSOLUTE_PATH+r"/INPUT/")
    print('df_agende_unique_fe',df_agende_unique_fe)
    return df_cluster,df_reclustering,df_cluster,df_cup_eapoc,calendar_fe,scheduler_a01,scheduler_a02,scheduler_a03,scheduler_a04,dict_peso_prestazioni_tutto,df_agende_unique_fe

## IMPORT FILE DA GENERAZIONE ANALISI SMART SCHEDULER 
df_cluster,df_reclustering,df_cluster,df_cup_eapoc,calendar_fe,scheduler_a01,scheduler_a02,scheduler_a03,scheduler_a04,dict_peso_prestazioni_tutto,df_agende_unique_fe = import_funzione_fe()
end_time = time.time()
execution_time = end_time - start_time
print(f"Il tempo di esecuzione è: {execution_time} secondi")

## FUNZIONI PER CALCOLO DF CALCOLATI ALL'INTERNO DELLA DASHBOARD 

def create_report_download(data_inizio,data_fine,lista_agende,recap_post_riconfigurazione_download):
    #print("________________PRINT CHECK______________")
    #print(lista_agende)
    #print(recap_post_riconfigurazione_download)
    #print(data_inizio)
    #print(data_fine)
    #print(recap_post_riconfigurazione_download.columns)
    data_inizio = pd.to_datetime(data_inizio)
    data_fine = pd.to_datetime(data_fine)
    priorita_filt = priorita[['STS11','CODICE_EROGANTE','DENOMINAZIONE_STRUTTURA','CODICE_AGENDA','DESCRIZIONE_AGENDA','UNITA_EROGANTE']].drop_duplicates()
    priorita_filt['CODICE_AGENDA'] = priorita_filt['CODICE_AGENDA'].astype(str)
    priorita_filt['CODICE_EROGANTE'] = priorita_filt['CODICE_EROGANTE'].astype(str)

    calendar_agende_filt = calendar_agende[['CALENDAR_DATE','DIARY_ID']].drop_duplicates()  ##'TIMEBAND_ID',
    calendar_agende_filt.rename(columns = {'DIARY_ID' : 'CODICE_AGENDA'},inplace=True)
    calendar_agende_filt['CALENDAR_DATE'] = pd.to_datetime(calendar_agende_filt['CALENDAR_DATE'])
    calendar_agende_filt_period = calendar_agende_filt[(calendar_agende_filt['CALENDAR_DATE'] >= data_inizio) & (calendar_agende_filt['CALENDAR_DATE'] <= data_fine)]


    recap_post_riconfigurazione_download['Giorno'] = pd.to_datetime(recap_post_riconfigurazione_download['Giorno'])
    
    recap_post_riconfigurazione_filt = recap_post_riconfigurazione_download[['Giorno','Codice_agenda','Slot_assegnati_U', 'Slot_assegnati_B','Slot_assegnati_D', 'Slot_assegnati_P']]
    recap_post_riconfigurazione_filt.rename(columns = {'Codice_agenda' : 'CODICE_AGENDA'},inplace=True)
    recap_post_riconfigurazione_filt['CODICE_AGENDA'] =recap_post_riconfigurazione_filt['CODICE_AGENDA'].astype(str)


    ## merge agende_attive--->calendar
    calendar_agende_attive = calendar_agende_filt_period.merge(priorita_filt, on ='CODICE_AGENDA', how ='left')
    calendar_agende_attive['CODICE_AGENDA'] = calendar_agende_attive['CODICE_AGENDA'].astype(str)


    ## lista valori 
    calendar_agende_filt = calendar_agende[['CALENDAR_DATE','DIARY_ID','TIMEBAND_ID']].drop_duplicates()  ##'TIMEBAND_ID',
    lista_timeband = calendar_agende_filt.groupby(['CALENDAR_DATE','DIARY_ID'])['TIMEBAND_ID'].unique().reset_index(name='TIMEBAND_ID')
    lista_timeband['CALENDAR_DATE'] = pd.to_datetime(lista_timeband['CALENDAR_DATE'])
    lista_timeband.rename(columns = {'DIARY_ID' : 'CODICE_AGENDA'}, inplace=True)


    #merge calendat_agende_attive con recap 
    calendar_agende_attive_recap = calendar_agende_attive.merge(recap_post_riconfigurazione_filt,left_on =['CALENDAR_DATE','CODICE_AGENDA'], right_on = ['Giorno','CODICE_AGENDA'], how='inner')

    ## MERGE TIMEBAND
    calendar_agende_attive_recap_final = calendar_agende_attive_recap.merge(lista_timeband, on = ['CALENDAR_DATE','CODICE_AGENDA'], how ='left')
    calendar_agende_attive_recap_final

    
    final_report_dettaglio = calendar_agende_attive_recap_final[['CALENDAR_DATE','STS11','DENOMINAZIONE_STRUTTURA','CODICE_EROGANTE','UNITA_EROGANTE','CODICE_AGENDA','DESCRIZIONE_AGENDA','TIMEBAND_ID','Slot_assegnati_U','Slot_assegnati_B','Slot_assegnati_D','Slot_assegnati_P']]
    final_report_dettaglio['CALENDAR_DATE'] = final_report_dettaglio['CALENDAR_DATE'].dt.strftime('%d/%m/%Y')
    final_report_dettaglio['TIMEBAND_ID'] = final_report_dettaglio['TIMEBAND_ID'].astype(str)
    
    final_report_dettaglio_filt_agende = final_report_dettaglio[final_report_dettaglio['CODICE_AGENDA'].isin(lista_agende)]
    
    final_sintetico = final_report_dettaglio_filt_agende.groupby(['STS11','DENOMINAZIONE_STRUTTURA','CODICE_EROGANTE','UNITA_EROGANTE','CODICE_AGENDA','DESCRIZIONE_AGENDA'])[['Slot_assegnati_U'	,'Slot_assegnati_B'	,'Slot_assegnati_D',	'Slot_assegnati_P']].sum().reset_index()
    final_sintetico['TOTALE'] = final_sintetico['Slot_assegnati_U'] + final_sintetico['Slot_assegnati_B'] + final_sintetico['Slot_assegnati_D'] + final_sintetico['Slot_assegnati_P']
    final_sintetico['PERC_U'] = ((final_sintetico['Slot_assegnati_U'] / final_sintetico['TOTALE']) * 100).round(2)
    final_sintetico['PERC_B'] = ((final_sintetico['Slot_assegnati_B'] / final_sintetico['TOTALE']) * 100).round(2)
    final_sintetico['PERC_D'] = ((final_sintetico['Slot_assegnati_D'] / final_sintetico['TOTALE']) * 100).round(2)
    final_sintetico['PERC_P'] = ((final_sintetico['Slot_assegnati_P'] / final_sintetico['TOTALE']) * 100).round(2)
    final_sintetico['data_inizio'] = pd.to_datetime(data_inizio)
    final_sintetico['data_inizio'] = final_sintetico['data_inizio'].dt.strftime('%d/%m/%Y')
    final_sintetico['data_fine'] = pd.to_datetime(data_fine)
    final_sintetico['data_fine'] = final_sintetico['data_fine'].dt.strftime('%d/%m/%Y')
    final_sintetico = final_sintetico[['data_inizio','data_fine','STS11', 'DENOMINAZIONE_STRUTTURA', 'CODICE_EROGANTE', 'UNITA_EROGANTE',
        'CODICE_AGENDA', 'DESCRIZIONE_AGENDA',
        'PERC_U', 'PERC_B', 'PERC_D', 'PERC_P']]
    
    
    return final_report_dettaglio_filt_agende,final_sintetico



## CALCOLO DELL'AMBITO AD OGNI CODICE AGENDA 
decodifica_ambiti_garanzia_rename = decodifica_ambiti_garanzia.rename(columns={'sts11': 'STS11', 'ambiti_aziende': 'AMBITI_AZIENDE'})
cols=['STS11', 'CODICE_EROGANTE', 'CODICE_AGENDA', 'SERVICE_REGIONAL_CODE', 'ORDER_DURATION', 'CLASSI_PRIORITA', 'AMBITI_AZIENDE']
df_agende_no_prestazioni_nuove = priorita[~(priorita['ORDER_ID'].str.endswith('_5', na=False))] 
df_agende_no_prestazioni_nuove_merge = df_agende_no_prestazioni_nuove.merge(decodifica_ambiti_garanzia_rename[['STS11', 'AMBITI_AZIENDE']], how='left', on='STS11')
df_agende_unique = df_agende_no_prestazioni_nuove_merge[cols].drop_duplicates()
df_agende_unique['tripletta'] = df_agende_unique['CODICE_AGENDA'].astype(str) + '_' + df_agende_unique['STS11'] + '_' + df_agende_unique['CODICE_EROGANTE'].astype(str)
decodifica_agende_ambito = df_agende_unique.groupby(['CODICE_AGENDA','STS11','CODICE_EROGANTE','AMBITI_AZIENDE']).first().reset_index()[['CODICE_AGENDA','STS11','CODICE_EROGANTE','AMBITI_AZIENDE']]



merged_df = pd.merge(priorita_nuovo, decodifica_agende_ambito, 
                     on=['CODICE_AGENDA', 'CODICE_EROGANTE', 'STS11'], 
                     how='inner')

decodifica_agende_ambito.rename(columns = {'CODICE_AGENDA' : 'Codice Agenda'}, inplace=True)

## FUNZIONI NETWORK E SUGGERIMENTO 

def suggerimento_intelligente(): 
    merge_2 = prestazioni_agende_df.merge(cluster[['Codice Prestazione','Comunità']], on ='Codice Prestazione', how ='left')
    ### funzione suggerimento 
    all_differenze = pd.DataFrame(columns=['Codice Agenda', 'Prestazioni'])
    lista_unica = merge_2['Codice Agenda'].unique()
    for cod in lista_unica:
        filtered_1 = merge_2[merge_2['Codice Agenda'] == cod]
        lista_community = filtered_1['Comunità'].unique()
        #print('lista' , lista_community)
        community_filt = cluster[cluster['Comunità'].isin(lista_community)]
        set_prestazioni_community = set(community_filt['Codice Prestazione'])
        #print('set_prestazioni_community',set_prestazioni_community)
        set_prestazioni_agenda = set(filtered_1['Codice Prestazione'])
        #print('set_prestazioni_agenda',set_prestazioni_agenda)
        differenza = set_prestazioni_community -set_prestazioni_agenda
        differenza_list = list(differenza)
        differenza_df = pd.DataFrame(differenza_list, columns=['Prestazioni'])
        differenza_df['Codice Agenda'] = cod
        # Concateniamo al DataFrame finale
        all_differenze = pd.concat([all_differenze, differenza_df], ignore_index=True)
    decodifica_catalogo.rename(columns = {'COD_PRESTAZIONE_CATALOGO' : 'Prestazioni'},inplace=True) 
    df_suggerimento = all_differenze.merge(decodifica_catalogo, on ='Prestazioni', how ='left')
    df_suggerimento.rename(columns = {'DESC_PRESTAZIONE_CATALOGO': 'Descrizione prestazione'},inplace=True)
    df_suggerimento.rename(columns = {'Prestazioni': 'Codice prestazione'},inplace=True)
    return df_suggerimento


def statistiche_aderenza_al_grafo(raggruppam):
    ## Funzione per creare DF --> 3 aderenze al grafo e df_chart che sarebbe il donut-chart --> Pagina Ulteriori dettagli
    prestazioni_agende_df['Codice Agenda'] = prestazioni_agende_df['Codice Agenda'].astype('str')
    # FILTRO A GRAFO PER RENDERLO SIGNIFICATIVO

    edge_list_completa_22_23_filt = edge_list_completa_22_23[edge_list_completa_22_23['cnt'] > 10]

    # PULISCI AGENDE DAI CODICI NON IN CATALOGO E FAI UN SELF JOIN 

    filt_new_prest = prestazioni_agende_df[~prestazioni_agende_df['Codice Prestazione'].str.contains('_')]
    filt_new_prest['len'] = filt_new_prest['Codice Prestazione'].str.len()
    filt_new_prest = filt_new_prest[filt_new_prest['len'] == 9]
    filt_new_prest.drop(['len'], axis= 1, inplace=True)
    agende_self = filt_new_prest.merge(filt_new_prest, on='Codice Agenda', how='left', suffixes=('', '_opp'))
    agende_self = agende_self[agende_self['Codice Prestazione'] > agende_self['Codice Prestazione_opp']]
    agende_self = agende_self[[ 'Codice Agenda', 'Codice Prestazione',
        'Descrizione Prestazione', 'Codice Prestazione_opp',
        'Descrizione Prestazione_opp']]
    edge_list_completa_22_23_filt.rename(columns = {'CODICE_PRESTAZIONE_CATALOGO' : 'Codice Prestazione', 'CODICE_PRESTAZIONE_CATALOGO_opp' : 'Codice Prestazione_opp'}, inplace=True)
    aderenza = agende_self.merge(edge_list_completa_22_23_filt, on = ['Codice Prestazione', 'Codice Prestazione_opp'],  how='outer', indicator=True)
    aderenza['Aderenza'] = aderenza.apply(lambda x: 'Si' if x['_merge'] == 'both' else 'No', axis=1)

    somma_tot = aderenza.groupby('Codice Agenda').size().reset_index(name ='somma_tot')

    aderenza_si = aderenza[aderenza['Aderenza'] == 'Si']
    somma_tot_aderenza_si = aderenza_si.groupby('Codice Agenda').size().reset_index(name ='somma_tot_si')
    finale_somma = somma_tot.merge(somma_tot_aderenza_si, on ='Codice Agenda', how ='left')
    finale_somma['Congruenza al grafo'] = round(finale_somma['somma_tot_si'] /finale_somma['somma_tot'],2)
    finale_somma.fillna(0,inplace=True)
    finale_somma = finale_somma[['Codice Agenda','Congruenza al grafo']]
    # RIMUOVIAMO TUTTI I CODICI NON CONFORMI -> QUELLI NUOVI E CON DIMESNIONE DIVERSA DA 9 
    filt_new_prest = prestazioni_agende_df[~prestazioni_agende_df['Codice Prestazione'].str.contains('_')]
    filt_new_prest['len'] = filt_new_prest['Codice Prestazione'].str.len()
    filt_new_prest = filt_new_prest[filt_new_prest['len'] == 9]
    filt_new_prest.drop(['len'], axis= 1, inplace=True)
    agende_self = filt_new_prest.merge(filt_new_prest, on='Codice Agenda', how='left', suffixes=('', '_opp'))
    agende_self = agende_self[agende_self['Codice Prestazione'] > agende_self['Codice Prestazione_opp']]
    agende_self = agende_self[[ 'Codice Agenda', 'Codice Prestazione',
        'Descrizione Prestazione', 'Codice Prestazione_opp',
        'Descrizione Prestazione_opp']]


    cluster_filt = cluster[['Codice Prestazione', 'Comunità']]
    ## SELF JOIN CLUSTER PER CONFRONTARE LE EDGE LIST 
    cluster_self_merge = cluster_filt.merge(cluster_filt, on='Comunità', how='left', suffixes=('', '_opp'))
    cluster_self_merge_filt = cluster_self_merge[cluster_self_merge['Codice Prestazione'] > cluster_self_merge['Codice Prestazione_opp']]
    ## MERGE DI AGENDE E CLUSTER 
    merge_agende_cluster = agende_self.merge(cluster_self_merge_filt,on = ['Codice Prestazione','Codice Prestazione_opp'], how = 'outer',indicator =True)
    somma_tot = agende_self.groupby('Codice Agenda').size().reset_index(name ='cnt_tot')
    merge_agende_cluster['Aderenza'] = merge_agende_cluster.apply(lambda x: 'Si' if x['_merge'] == 'both' else 'No', axis=1)
    merge_agenda_cluster_only_si = merge_agende_cluster[merge_agende_cluster['Aderenza'] == 'Si']
    somma_tot_si = merge_agenda_cluster_only_si.groupby('Codice Agenda').size().reset_index(name ='cnt_tot_si')
    confronto_perc_cluster = somma_tot.merge(somma_tot_si , on ='Codice Agenda', how ='left')
    confronto_perc_cluster['Congruenza al cluster'] =round(confronto_perc_cluster['cnt_tot_si'] / confronto_perc_cluster['cnt_tot'],2)
    confronto_perc_cluster.fillna(0,inplace=True)
    confronto_perc_cluster['Codice Agenda'] = confronto_perc_cluster['Codice Agenda'].astype('str')
    confronto_perc_cluster = confronto_perc_cluster[['Codice Agenda','Congruenza al cluster']]

    ## PULIZIA DEI CODICI NON CONFORMI SECODNO SOLO 2 REGOLE PER ORA PRESTAZIONE NON CONTIENE '_' E NON è DIVERSA DA 9 CARATTERI DI LUGNHEZZA
    raggruppam['CODICE_REGIONALE_PREST'] = raggruppam['CODICE_REGIONALE_PREST'].astype('str')
    raggruppam = raggruppam[~raggruppam['CODICE_REGIONALE_PREST'].str.contains('_')]
    raggruppam['len'] = raggruppam['CODICE_REGIONALE_PREST'].str.len()
    raggruppam_filt = raggruppam[raggruppam['len'] == 9]
    raggruppam_filt.drop(['len'], axis= 1, inplace=True)
    raggruppam_filt['CODICE_AGENDA'] = raggruppam_filt['CODICE_AGENDA'].astype('str')
    raggruppam_filt_col = raggruppam_filt[['COD_AZIENDA_x','DESCRIZIONE_AZIENDA_x','TIPO_AZIENDA_x','STS11','CODICE_AGENDA','DESCRIZIONE_AGENDA_x','RAGGRUPPAMENTO','CODICE_REGIONALE_PREST','DESCR_PREST_y']]
    self_merge = raggruppam_filt_col.merge(raggruppam_filt_col, on =['RAGGRUPPAMENTO','CODICE_AGENDA'])
    self_merge_edge = self_merge[self_merge['CODICE_REGIONALE_PREST_x'] > self_merge['CODICE_REGIONALE_PREST_y']]
    self_merge_edge.rename(columns = {'CODICE_REGIONALE_PREST_x' :  'CODICE_PRESTAZIONE_CATALOGO' ,'CODICE_REGIONALE_PREST_y' :'CODICE_PRESTAZIONE_CATALOGO_opp'},inplace=True)

    edge_22_23 = edge_list_completa_22_23.copy()
    edge_22_23 = edge_22_23[edge_22_23['cnt'] >10]
    self_merge_edge['CODICE_PRESTAZIONE_CATALOGO'] = self_merge_edge['CODICE_PRESTAZIONE_CATALOGO'].astype('str')
    self_merge_edge['CODICE_PRESTAZIONE_CATALOGO_opp'] = self_merge_edge['CODICE_PRESTAZIONE_CATALOGO_opp'].astype('str')
    edge_22_23['CODICE_PRESTAZIONE_CATALOGO'] = edge_22_23['CODICE_PRESTAZIONE_CATALOGO'].astype('str')
    edge_22_23['CODICE_PRESTAZIONE_CATALOGO_opp'] = edge_22_23['CODICE_PRESTAZIONE_CATALOGO_opp'].astype('str')

    ragguppam_finale = self_merge_edge.merge(edge_22_23, on =['CODICE_PRESTAZIONE_CATALOGO','CODICE_PRESTAZIONE_CATALOGO_opp'], how = 'outer',indicator =True)

    somma_tot = self_merge_edge.groupby('CODICE_AGENDA').size().reset_index(name ='cnt_tot')
    ragguppam_finale['Aderenza'] = ragguppam_finale.apply(lambda x: 'Si' if x['_merge'] == 'both' else 'No', axis=1)
    merge_agenda_cluster_only_si = ragguppam_finale[ragguppam_finale['Aderenza'] == 'Si']
    somma_tot_si = merge_agenda_cluster_only_si.groupby('CODICE_AGENDA').size().reset_index(name ='cnt_tot_si')
    confronto_perc_raggru = somma_tot.merge(somma_tot_si , on ='CODICE_AGENDA', how ='left')
    confronto_perc_raggru['Congruenza ai raggruppamenti'] =round(confronto_perc_raggru['cnt_tot_si'] / confronto_perc_raggru['cnt_tot'],2)
    confronto_perc_raggru.fillna(0,inplace=True)
    confronto_perc_raggru.rename(columns ={'CODICE_AGENDA' : 'Codice Agenda'},inplace=True)
    confronto_perc_raggru = confronto_perc_raggru[['Codice Agenda','Congruenza ai raggruppamenti']]
    df_dettaglio_agenda_1 = confronto_perc_raggru.merge(confronto_perc_cluster, on ='Codice Agenda', how ='outer' )
    df_dettaglio_agenda_1 = df_dettaglio_agenda_1.merge(finale_somma, on ='Codice Agenda', how ='outer' )
    ## DF-CHART
    def calculate_weighted_value(row):
        # Define weights
        weights = {
            'Congruenza al grafo': 0.25,
            'Congruenza al cluster': 0.35,
            'Congruenza ai raggruppamenti': 0.45
        }
        
        # Get non-null columns and their corresponding weights
        non_null_columns = [col for col in weights if pd.notnull(row[col])]
        total_weight = sum(weights[col] for col in non_null_columns)
        
        # Normalize weights for non-null columns
        normalized_weights = {col: weights[col] / total_weight for col in non_null_columns}
        
        # Calculate the weighted value
        value = sum(row[col] * normalized_weights[col] for col in non_null_columns)
        return value
    df_dettaglio_agenda_aderenza = df_dettaglio_agenda_1.copy()
    df_dettaglio_agenda_aderenza['Value'] = round(df_dettaglio_agenda_1.apply(calculate_weighted_value, axis=1),2)
    agende_grado_aderenza = df_dettaglio_agenda_aderenza[['Codice Agenda','Value']]


    return agende_grado_aderenza,df_dettaglio_agenda_1


###########################################
## DF CALCOLATI DA FUNZIONI CREATE  #######
##########################################


df_chart,df_dettaglio_agenda_1 = statistiche_aderenza_al_grafo(raggruppam)
df_suggerimento = suggerimento_intelligente()



### CAPIRE SE SI PUO ELIMINARE 
dict_ambiti = [
    {'label': 'Provincia di Caserta', 'value': 'ASL Caserta A04'},
    {'label': 'Provincia di Caserta 1', 'value': 'ASL Caserta 1'},
    {'label': 'Provincia di Caserta 2', 'value': 'ASL Caserta 2'},
    {'label': 'Provincia di Caserta 3', 'value': 'ASL Caserta 3'}
]

# Creare un dizionario che mappa il value alla label corrispondente
value_to_label = {item['value']: item['label'] for item in dict_ambiti}
vdati = {
    "Stato attuale": [
        {"Priorità": "Urgente", "Immagine": "urgent.png", "Slot": "-", "Giorni": "-", "Slot_Riconfigurazione": "-", "Giorni_Riconfigurazione": "-"},
        {"Priorità": "Breve", "Immagine": "short.png", "Slot": "-", "Giorni": "-", "Slot_Riconfigurazione": "-", "Giorni_Riconfigurazione": "-"},
        {"Priorità": "Differibile", "Immagine": "deferable.png", "Slot": "-", "Giorni": "-", "Slot_Riconfigurazione": "-", "Giorni_Riconfigurazione": "-"},
        {"Priorità": "Programmata", "Immagine": "planned.png", "Slot": "-", "Giorni": "-", "Slot_Riconfigurazione": "-", "Giorni_Riconfigurazione": "-"},
    ],
    
    "Stato post riconfigurazione": [
        {"Priorità": "Urgente", "Immagine": "urgent.png", "Slot": "-", "Giorni": "-", "Slot_Riconfigurazione": "-", "Giorni_Riconfigurazione": "-"},
        {"Priorità": "Breve", "Immagine": "short.png", "Slot": "-", "Giorni": "-", "Slot_Riconfigurazione": "-", "Giorni_Riconfigurazione": "-"},
        {"Priorità": "Differibile", "Immagine": "deferable.png", "Slot": "-", "Giorni": "-", "Slot_Riconfigurazione": "-", "Giorni_Riconfigurazione": "-"},
        {"Priorità": "Programmata", "Immagine": "planned.png", "Slot": "-", "Giorni": "-", "Slot_Riconfigurazione": "-", "Giorni_Riconfigurazione": "-"},
    ]
}

dati = {
    "Stato attuale": [
        {"Priorità": "Urgente", "Immagine": "urgent.png", "Slot": "-", "Giorni": "-", "Slot_Riconfigurazione": "-", "Giorni_Riconfigurazione": "-"},
        {"Priorità": "Breve", "Immagine": "short.png", "Slot": "-", "Giorni": "-", "Slot_Riconfigurazione": "-", "Giorni_Riconfigurazione": "-"},
        {"Priorità": "Differibile", "Immagine": "deferable.png", "Slot": "-", "Giorni": "-", "Slot_Riconfigurazione": "-", "Giorni_Riconfigurazione": "-"},
        {"Priorità": "Programmata", "Immagine": "planned.png", "Slot": "-", "Giorni": "-", "Slot_Riconfigurazione": "-", "Giorni_Riconfigurazione": "-"},
    ],
    
    "Stato post riconfigurazione": [
        {"Priorità": "Urgente", "Immagine": "urgent.png", "Slot": "-", "Giorni": "-", "Slot_Riconfigurazione": "-", "Giorni_Riconfigurazione": "-"},
        {"Priorità": "Breve", "Immagine": "short.png", "Slot": "-", "Giorni": "-", "Slot_Riconfigurazione": "-", "Giorni_Riconfigurazione": "-"},
        {"Priorità": "Differibile", "Immagine": "deferable.png", "Slot": "-", "Giorni": "-", "Slot_Riconfigurazione": "-", "Giorni_Riconfigurazione": "-"},
        {"Priorità": "Programmata", "Immagine": "planned.png", "Slot": "-", "Giorni": "-", "Slot_Riconfigurazione": "-", "Giorni_Riconfigurazione": "-"},
    ]
}








#df_data = pd.read_csv(root_path+r"\prima_data_disponibile.csv",sep =',')
# Funzione per formattare i numeri con il punto come separatore delle migliaia
def format_thousands(x):
    try:
        return "{:,.0f}".format(x).replace(",", ".")
    except (ValueError, TypeError):
        return x

# Applica la formattazione a tutte le colonne numeriche
#df_slot_giorni = df_slot_giorni.applymap(lambda x: format_thousands(x) if isinstance(x, (int, float)) else x)

# APP.LAYOUT

app = dash.Dash(__name__, external_stylesheets=[dbc.themes.BOOTSTRAP])
server = app.server  # the Flask instance used by Dash
app.layout = html.Div([
    dcc.Location(id='url', refresh=False),
    dcc.Location(id='url-redirect', refresh=True),  # Added url-redirect
    #dcc.Store(id='calendar-filtered'),
    dcc.Store(id='stored-filename',storage_type='session'),
    dcc.Store(id='stored-contents',storage_type='session'),
    dcc.Store(id='start-date-store',storage_type='session'),
    dcc.Store(id='end-date-store',storage_type='session'),
    dcc.Store(id='start-day-store',storage_type='session'),
    dcc.Store(id='start-month-store',storage_type='session'),
    dcc.Store(id='start-year-store',storage_type='session'),
    dcc.Store(id='end-day-store',storage_type='session'),
    dcc.Store(id='end-month-store',storage_type='session'),
    dcc.Store(id='end-year-store',storage_type='session'),
    dcc.Store(id='selected-code-store',storage_type='session'),   ## Servire a tenere traccia delle righe selezionate 
    dcc.Store(id='dropdown-value-store',storage_type='session'),  # Nuovo Store per il valore del dropdown 
    dcc.Store(id='username1-store',storage_type='session'),
    dcc.Store(id='button-1-clicks-store',storage_type='session'),
    dcc.Store(id='username2-store',storage_type='session'),
    dcc.Store(id='button-2-clicks-store',storage_type='session'),
    dcc.Store(id='selected-rows',storage_type='session'),
    dcc.Store(id = 'selected-ambiti-store',storage_type='session'),
    dcc.Store(id='store-df1',storage_type='session'),
    dcc.Store(id='store-df2',storage_type='session'),
    dcc.Store(id='store-df3',storage_type='session'),
    dcc.Store(id='store-df4',storage_type='session'),
    dcc.Store(id='store-df5',storage_type='session'),
    dcc.Store(id='store-df6',storage_type='session'),
    dcc.Store(id='data-ready', data=False),
    dcc.Store(id = 'initial-options',storage_type='session'),
    dcc.Store(id = 'name',storage_type='session'),
    html.Div(id='page-content')
])



### CALLBACK FUNZIONAMENTO DASHBOARD 


# Callback to update page content based on URL
@app.callback(
    Output('page-content', 'children'),
    [Input('url', 'pathname')],
)
def display_page(pathname):
    if pathname == '/agende':
        return layout_agende()
    elif pathname == '/tempi_attesa':
        return layout_tempi_attesa()
    elif pathname == '/ulteriori_dettagli':
        return layout_ulteriori_dettagli()
    elif pathname== '/calendario':
        return layout_home()
    elif pathname =='/consigli':
        return consigli_di_dimensionamento()
    else:
        return layout_home()
    
    
### callback per filtro tabella pagina 1
@app.callback(
    Output('selected-rows-container', 'children'),
    Input('selected-code-store', 'data')
)
def display_selected_rows(selected_codes):
    if not selected_codes:
        return "Nessuna riga selezionata."
    return html.Pre(f"Codici selezionati:\n{selected_codes}")
## Callback per aggiornare i dati dello store quando le righe vengono selezionate
@app.callback(
    Output('selected-code-store', 'data'),
    Input('table-new', 'selectedRows')
    )
def update_selected_code(selected_rows):
    if selected_rows:
        selected_codes = [row['Codice Agenda'] for row in selected_rows]
        return selected_codes
    return []

"""@app.callback(
    Output('selected-rows', 'data'),
    Input('table-new', 'selectedRows'),
    )
def update_selected_code(selected_rows):
   
    return selected_rows"""

@app.callback(
    Output('table-new', 'selectedRows'),
    [Input('selected-code-store', 'data'),
    Input('store-df6', 'data')],
    
    )
def update_selected_code(selected_codes,df_data_agende_dict):
    if selected_codes:
        data_agende = pd.DataFrame.from_dict(df_data_agende_dict)
        selected_rows = [row for row in data_agende.to_dict('records') if row['Codice Agenda'] in selected_codes]
        return selected_rows
    else:
        return []
    

####


# Aggiorna la callback per leggere e scrivere il valore selezionato dallo store
@app.callback(
    [Output('dropdown-tempi-attesa', 'options'),
     Output('dropdown-tempi-attesa', 'value')],
    [Input('url', 'pathname'), Input('selected-code-store', 'data'),Input('dropdown-value-store', 'data'),Input('store-df6', 'data')],
    
)
def update_dropdown(pathname, selected_codes, selected_value,df_data_agende_dict):
    if pathname == '/tempi_attesa' and selected_codes:
        data_agende = pd.DataFrame.from_dict(df_data_agende_dict)
        options = [{'label': data_agende.loc[data_agende['Codice Agenda'] == code, 'Descrizione Agenda'].values[0], 'value': code} for code in selected_codes]
        # Se c'è un valore selezionato memorizzato, usa quello
        if selected_value in selected_codes:
            selected_value = str(selected_value)
            return options, selected_value
        else:
            value = options[0]['value'] if options else None
            return options, value
    return [], None




@app.callback(
    [Output('table-actual-slot', 'children'),
     Output('table-actual-days', 'children'),],
    [Input('dropdown-tempi-attesa', 'value'), Input('store-df1','data')]
)
def update_tables(selected_code, four_tables_dict):
    df_slot_giorni = pd.DataFrame.from_dict(four_tables_dict) if four_tables_dict else pd.DataFrame()
    if selected_code:
        selected_code = str(selected_code)
        filtered_data = df_slot_giorni[df_slot_giorni['Codice Agenda'] == selected_code]
        
        def update_data(original_data, filtered_data, slot_col, giorni_col):
            updated_data = []
            giorni_values = []

            for row in original_data:
                if row["Priorità"] in filtered_data['Priorità'].values:
                    row_data = filtered_data[filtered_data['Priorità'] == row["Priorità"]].iloc[0]
                    updated_row = row.copy()
                    updated_row["Slot"] = row_data[slot_col] if not pd.isna(row_data[slot_col]) else "-"
                    updated_row["Giorni"] = row_data[giorni_col] if not pd.isna(row_data[giorni_col]) else "-"
                    
                    if updated_row["Giorni"] != "-" and updated_row["Giorni"] != "xx giorni":
                        giorni_values.append(int(updated_row["Giorni"]))

                    updated_data.append(updated_row)
                else:
                    updated_row = row.copy()
                    updated_row["Slot"] = "-"
                    updated_row["Giorni"] = "-"
                    updated_data.append(updated_row)

            return updated_data
        
        # Update data per ogni sezione
        updated_stato_attuale_slot = update_data(dati["Stato attuale"], filtered_data, "Slot_Attuale", "Giorni_Attuale")
        updated_stato_post_riconfigurazione_slot = update_data(dati["Stato post riconfigurazione"], filtered_data, "Slot_Riconfigurazione", "Giorni_Riconfigurazione")
        updated_stato_attuale_days = update_data(dati["Stato attuale"], filtered_data, "Slot_Attuale", "Giorni_Attuale")
        updated_stato_post_riconfigurazione_days = update_data(dati["Stato post riconfigurazione"], filtered_data, "Slot_Riconfigurazione", "Giorni_Riconfigurazione")
                
        # Applica la formattazione ai campi Slot e Giorni
        for updated_row in updated_stato_attuale_slot:
            if updated_row["Slot"] != "-":
                updated_row["Slot"] = format_thousands(updated_row["Slot"])
            if updated_row["Giorni"] != "-" and updated_row["Giorni"] != "xx giorni":
                updated_row["Giorni"] = format_thousands(updated_row["Giorni"])

        for updated_row in updated_stato_post_riconfigurazione_slot:
            if updated_row["Slot"] != "-":
                updated_row["Slot"] = format_thousands(updated_row["Slot"])
            if updated_row["Giorni"] != "-" and updated_row["Giorni"] != "xx giorni":
                updated_row["Giorni"] = format_thousands(updated_row["Giorni"])

        # Creazione di una nuova lista di dizionari con la colonna aggiuntiva Slot_Riconfigurazione
        updated_stato_attuale_slot_updated = []

        for updated_row_attuale, updated_row_riconfigurazione in zip(updated_stato_attuale_slot, updated_stato_post_riconfigurazione_slot):
            updated_row_attuale_updated = updated_row_attuale.copy()  # Copia il dizionario esistente
            updated_row_attuale_updated["Slot_Riconfigurazione"] = updated_row_riconfigurazione["Slot"]  # Aggiungi la nuova colonna
            updated_stato_attuale_slot_updated.append(updated_row_attuale_updated)
        
        for updated_row in updated_stato_attuale_days:
            if updated_row["Slot"] != "-":
                updated_row["Slot"] = format_thousands(updated_row["Slot"])
            if updated_row["Giorni"] != "-" and updated_row["Giorni"] != "xx giorni":
                updated_row["Giorni"] = format_thousands(updated_row["Giorni"])

        for updated_row in updated_stato_post_riconfigurazione_days:
            if updated_row["Slot"] != "-":
                updated_row["Slot"] = format_thousands(updated_row["Slot"])
            if updated_row["Giorni"] != "-" and updated_row["Giorni"] != "xx giorni":
                updated_row["Giorni"] = format_thousands(updated_row["Giorni"])

        # Creazione di una nuova lista di dizionari con la colonna aggiuntiva Giorni_Riconfigurazione
        updated_stato_attuale_days_updated = []

        for updated_row_attuale, updated_row_riconfigurazione in zip(updated_stato_attuale_days, updated_stato_post_riconfigurazione_days):
            updated_row_attuale_updated = updated_row_attuale.copy()  # Copia il dizionario esistente
            updated_row_attuale_updated["Giorni_Riconfigurazione"] = updated_row_riconfigurazione["Giorni"]  # Aggiungi la nuova colonna
            updated_stato_attuale_days_updated.append(updated_row_attuale_updated)

        return (
            create_table_rows(updated_stato_attuale_slot_updated, include_slot=True),
            create_table_rows(updated_stato_attuale_days_updated, include_days=True),
        )
    else:
        return (
            None, None
        )

def create_table_rows(data, include_days=False, include_slot=False, include_image=True):
    rows = []
    has_average_row = any(row['Priorità'] == "" for row in data)

    for row in data:
        class_name = ""
        if row["Immagine"] == "urgent.png":
            class_name = "urgent"
        elif row["Immagine"] == "short.png":
            class_name = "short"
        elif row["Immagine"] == "deferable.png":
            class_name = "deferable"
        elif row["Immagine"] == "planned.png":
            class_name = "planned"

        if include_image:
            row_cells = [
                html.Td(row["Priorità"], className="table-cell", style={"border-bottom": "none"}),
                html.Td(html.Img(src=f"/assets/{row['Immagine']}", style={"width": "30px"}), className=f"table-cell {class_name}", style={"border-bottom": "none"}) if row["Immagine"] else html.Td("", className="table-cell", style={"border-bottom": "none"})
            ]
        else:
            row_cells = [
                html.Td(row["Priorità"], className="table-cell", style={"border-bottom": "none"}, hidden=True),
                html.Td(html.Img(src=f"/assets/{row['Immagine']}", style={"width": "30px"}), className=f"table-cell {class_name}", style={"border-bottom": "none"}, hidden=True) if row["Immagine"] else html.Td("", className="table-cell", style={"border-bottom": "none"}, hidden=True)
            ]

        if include_slot:
            row_cells.append(html.Td(row["Slot"], className="table-cell", style={"border-bottom": "none"}))
            if "Slot_Riconfigurazione" in row:  # Aggiungi Slot_Riconfigurazione solo se presente
                row_cells.append(html.Td(row["Slot_Riconfigurazione"], className="table-cell", style={"border-bottom": "none"}))

        if include_days:
            row_cells.append(html.Td(row["Giorni"], className="table-cell", style={"border-bottom": "none"}))
            if "Giorni_Riconfigurazione" in row:  # Aggiungi Giorni_Riconfigurazione solo se presente
                row_cells.append(html.Td(row["Giorni_Riconfigurazione"], className="table-cell", style={"border-bottom": "none"}))

        # Linea blu per Media
        if has_average_row and row["Priorità"] == "":
            rows.append(html.Tr(html.Td(html.Hr(style={"border": "2px solid", "width": "100%", 'color': '#0F2471'}), colSpan=len(row_cells), style={"padding": "0", "border-bottom": "none"})))

        rows.append(html.Tr(row_cells))

    return rows



## callback per menu a tendina in alto ulteriori dettagli 
@app.callback(
    Output('agenda_name', 'children'),
    [Input('dropdown-value-store', 'data'),
    Input('store-df6', 'data')]
)
def update_agenda_viewer(selected_agenda,df_data_agende_dict):
    if selected_agenda:
        selected_agenda = str(selected_agenda)
        data_agende = pd.DataFrame.from_dict(df_data_agende_dict)
        if selected_agenda in data_agende['Codice Agenda'].values:
            description = data_agende.loc[data_agende['Codice Agenda'] == selected_agenda, 'Descrizione Agenda'].values[0]
            return html.Div(description, style={'white-space': 'pre-line'})
        else:
            return html.Span("Codice agenda non trovato.")
    return html.Span("Seleziona un codice agenda dal menu a tendina.")

#### Informazioni sulla struttura 

@app.callback(
    Output('output-container', 'children'),
    [Input('dropdown-value-store', 'data')]
)
def update_output(selected_agenda):
    filtered_df = df_dettaglio_struttura[df_dettaglio_struttura['Codice Agenda'].astype(str) == selected_agenda]
    if filtered_df.empty:
        return html.P('No data available')
    row = filtered_df.iloc[0]
    return html.Div([
        html.P([
            html.Div('Azienda: ', style={'fontWeight': 'bold', 'font-family': font_family_variable, 'color': '#0F2471'}),
            html.Div(row['Azienda'])
        ]),
        html.P([
            html.Div('Codice Erogante: ', style={'fontWeight': 'bold', 'font-family': font_family_variable, 'color': '#0F2471'}),
            html.Div(row['Codice Erogante'])
        ]),
        html.P([
            html.Div('Specializzazione: ', style={'fontWeight': 'bold', 'font-family': font_family_variable, 'color': '#0F2471'}),
            html.Div(row['Specializzazione'])
        ]),
    ])


# Callback per memorizzare il valore del dropdown selezionato
@app.callback(
    Output('dropdown-value-store', 'data'),
    [Input('dropdown-tempi-attesa', 'value')]
)
def store_dropdown_value(selected_code):
    return selected_code





# Callback per aggiornare la tabella filtrata
@app.callback(
    [Output('filtered-table-ulteriori', 'rowData'),
     Output('filtered-table-ulteriori', 'columnDefs'),
     Output('filtered-table-ulteriori', 'style')],
    [Input('dropdown-value-store', 'data')]
)
def update_filtered_table(selected_code):
    if selected_code:
        selected_code = str(selected_code)
        filtered_data = prestazioni_agende_df[prestazioni_agende_df['Codice Agenda'] == selected_code]
        filtered_data = filtered_data.drop('Codice Agenda', axis= 1)
        if not filtered_data.empty:
            columns = [{'headerName': i, 'field': i} for i in filtered_data.columns]
            data = filtered_data.to_dict('records')
            # Calcoliamo l'altezza della tabella in base al numero di righe
            table_height = 100 # Altezza minima di 40px, massima di 500px
            table_style = {'height': f'{table_height}%'}
            return data, columns, table_style
    # Se non ci sono dati, restituisci liste vuote per righe e colonne e stile della tabella vuoto
    return [], [], {'height': '0px'}


@app.callback(
    Output('chart-text', 'children'),
    [Input('dropdown-value-store', 'data')]
)
def update_donut_chart(selected_code):
    if selected_code:
        selected_code = str(selected_code)
        filtered_chart_data = df_chart[df_chart['Codice Agenda'] == selected_code]
        
        if not filtered_chart_data.empty:
            valore_aderenza = filtered_chart_data['Value'].values[0]

            values = [valore_aderenza * 100, (1 - valore_aderenza) * 100]
            labels = ['Grado di Cooperazione', ' ']
            
            # Creiamo il grafico a torta 3D
            fig = go.Figure(data=[go.Pie(
                labels=labels,
                values=values,
                hole=0.7,
                marker=dict(
                    colors=['#00338D', '#8aa0e0'],
                    line=dict(color='white', width=3)
                ),
                pull=[0.1],
                textinfo='label+percent',
                textposition='outside',
                insidetextfont=dict(
                    color=['#00338D', '#ECF2FE']  # Colori del testo interno per ogni label
                ),
                outsidetextfont=dict(
                    color=['#00338D', '#ECF2FE']  # Colori del testo esterno per ogni label
                )
            )])

            fig.update_traces(
                textposition='outside',
                textfont=dict(size=16, color='#0F2471', family='Arial', weight='bold')
            )

            fig.update_layout(
                paper_bgcolor='#ECF2FE',
                plot_bgcolor='#ECF2FE',
                showlegend=False,
                margin={'b': 0, 'l': 0, 't': 0, 'r': 0},
                height=None,  # Permette di essere responsivo
                width=None   # Permette di essere responsivo
            )

            return html.Div([
                dcc.Graph(figure=fig, style={'width': '100%', 'height': '100%'})
            ], style={'width': '100%', 'height': '90%'})  # Altezza massima del contenitore
        else:
            return html.Div("Nessun dato disponibile.")
    else:
        return html.Div("Seleziona un codice.")


"""
def update_donut_chart(selected_code):
    if selected_code:
        selected_code = str(selected_code)
        filtered_chart_data = df_chart[df_chart['Codice Agenda'] == selected_code]
        
        if not filtered_chart_data.empty:
            valore_aderenza = filtered_chart_data['Value'].values[0]

            values = [valore_aderenza, 1 - valore_aderenza]
            labels = ['Grado di Aderenza', 'Non Aderenza']
            
            fig = go.Figure(data=[go.Pie(
                labels=labels,
                values=values,
                hole=.7,
                marker=dict(colors=['#B5D0FF', 'lightgrey'], line=dict(color='white', width=3)),
                textinfo='none',
                hoverinfo='label+percent',
                hovertext=labels
            )])

            
            fig.update_layout(
                paper_bgcolor='#F2F2F2',
                plot_bgcolor='#F2F2F2',
                width=580,
                height=390,
                showlegend=False,
                annotations=[
                    dict(
                        text=f'{valore_aderenza*100:.0f}%', 
                        x=0.5, 
                        y=0.5, 
                        font=dict(color='#B5D0FF', family='Arial',weight='bold', size=24), 
                        showarrow=False
                    ),
                    dict(
                        text="Rappresenta la % di presenza nel<br>"
                             "medesimo cluster del grafo delle coppie di<br>"
                             "prestazioni in agenda.",
                        x=0.5,
                        y=-0.37,  # Spostiamo la ciambella un po' più in alto
                        align='center',
                        valign='middle',
                        font=dict(color='#00194C', size=15),
                        showarrow=False
                    )
                ]
            )

            return html.Div([
                dcc.Graph(figure=fig),
                #html.Div("Rappresentazione testuale aggiuntiva")
            ])

    # If no data is available or no code is selected, return a message
    return "Dati non presenti"
"""




### tab 2 
def create_pie_chart_with_annotation(labels, values, annotation_text):
    fig = go.Figure(go.Pie(labels=labels, values=values,hole=.7,textinfo='none'))
    fig.update_traces(marker=dict(colors=['#00338D', '#8aa0e0']))
    fig.update_layout(
        showlegend=False,
        paper_bgcolor='#ECF2FE',
        plot_bgcolor='#ECF2FE',
        annotations=[dict(text=annotation_text, x=0.5, y=0.5, font_size=20, showarrow=False)]
    )
    return fig

@app.callback(
    [Output('grafo-donut-chart', 'figure'),
     Output('cluster-donut-chart', 'figure'),
     Output('raggruppamenti-donut-chart', 'figure')],
    [Input('dropdown-value-store', 'data')]
)
def update_charts(selected_code):
    # Filtra il dataframe in base al codice selezionato
    selected_code = str(selected_code)
    df_filtered = df_dettaglio_agenda_1[df_dettaglio_agenda_1['Codice Agenda'] == selected_code]
    
    if not df_filtered.empty:
        grafo_value = df_filtered['Congruenza al grafo'].iat[0]
        cluster_value = df_filtered['Congruenza al cluster'].iat[0]
        raggruppamenti_value = df_filtered['Congruenza ai raggruppamenti'].iat[0]
    else:
        grafo_value = cluster_value = raggruppamenti_value = None
    
    # Create the donut charts
    # Grafico per "Congruenza al grafo"
    fig_grafo = create_pie_chart_with_annotation(
        labels=['Congruenza al grafo', ''],
        values=[grafo_value if grafo_value is not None else 0, 1 - grafo_value if grafo_value is not None else 1],
        annotation_text=f"{grafo_value*100:.0f}%" if pd.notna(grafo_value) else "-"
    )

    # Grafico per "Congruenza al cluster"
    fig_cluster = create_pie_chart_with_annotation(
        labels=['Congruenza al cluster', ''],
        values=[cluster_value if cluster_value is not None else 0, 1 - cluster_value if cluster_value is not None else 1],
        annotation_text=f"{cluster_value*100:.0f}%" if pd.notna(cluster_value) else "-"
    )

    # Grafico per "Congruenza ai raggruppamenti"
    fig_raggruppamenti = create_pie_chart_with_annotation(
        labels=['Congruenza ai raggruppamenti', ''],
        values=[raggruppamenti_value if raggruppamenti_value is not None else 0, 1 - raggruppamenti_value if raggruppamenti_value is not None else 1],
        annotation_text=f"{raggruppamenti_value*100:.0f}%" if pd.notna(raggruppamenti_value) else "-"
    )
    labels = ['Congruenza al grafo','Congruenza al cluster','Congruenza ai raggruppamenti']
    # Update traces and layout for each figure
    i=0
    for fig in [fig_grafo, fig_cluster, fig_raggruppamenti]:
        #for label in labels:
        fig.update_traces(marker=dict(colors=['#00338D','#8aa0e0']),textinfo='none')
        fig.update_layout(showlegend=False,paper_bgcolor='#ECF2FE',
                    plot_bgcolor='#ECF2FE',title_text = labels[i],title_font=dict(size=12, color='#00338D'),title=dict(x=0.50,yanchor='top',  # Anchor point for y
                y=0.95), margin=dict(b=10, l=10, r=10, t=35) )
        i+=1
    
    return fig_grafo, fig_cluster, fig_raggruppamenti


## filtro tabella suggerimento intellingente

@app.callback(
    [Output('table-suggerimento-intelligente', 'rowData'),
     Output('table-suggerimento-intelligente', 'columnDefs'),
     Output('table-suggerimento-intelligente', 'style')],
    [Input('dropdown-value-store', 'data')]
)
def update_table(selected_code):
    if selected_code:
        selected_code =str(selected_code)
        filtered_data = df_suggerimento[df_suggerimento['Codice Agenda'] == selected_code]
        filtered_data = filtered_data.drop('Codice Agenda', axis= 1)
        if not filtered_data.empty:
            columns = [{'headerName': i, 'field': i} for i in filtered_data.columns]
            data = filtered_data.to_dict('records')
            # Calcoliamo l'altezza della tabella in base al numero di righe
            table_height = 18 # Altezza minima di 40px, massima di 500px
            table_style = {'height': f'{table_height}vh'}
            return data, columns, table_style
    # Se non ci sono dati, restituisci liste vuote per righe e colonne e stile della tabella vuoto
    return [], [], {'height': '0px'}



# Example callback for download



# Callback for downloading the file
@app.callback(
    Output("download-file", "data"),
    [Input("download-image", "n_clicks"), Input('start-date-store','data'),Input('end-date-store','data'),Input('selected-code-store', 'data'),Input('store-df4','data')],
)
def update_download_data(n_clicks,start_date,end_date,options,recap_stato_post_riconfigurazione_dict):
 
    #print(df_report_percentuale)
    if n_clicks :
        
        df_daily_requests_in_minutes_per_agend = pd.read_csv(os.path.join(base_path,'File_utili','df_daily_requests_in_minutes_per_agend.csv'))
        #recap_stato_post_riconfigurazione_download = pd.DataFrame.from_dict(recap_stato_post_riconfigurazione_dict)
        #print('---recap download inner--', df_daily_requests_in_minutes_per_agend)
        # Path to the original file
        file_path = root_path_download

        # Load the existing Excel file
        workbook = load_workbook(file_path)
        current_date = datetime.now().strftime("%Y-%m-%d")
        download_filename = f"Smart Scheduler_Slot Giornalieri Consigliati_{current_date}.xlsx"
        df_report_dettaglio,df_report_percentuale = create_report_download(start_date,end_date,options,df_daily_requests_in_minutes_per_agend)
        workbook = load_workbook(file_path)
        #print('download')

        sheet_name_1 = 'Report_Dettaglio'
        sheet_name_2 = 'Report_Percentuale'
        worksheet_det = workbook[sheet_name_1]
        worksheet_per = workbook[sheet_name_2]



        start_row = 5
        start_col = 2


        for r_idx, row in enumerate(df_report_dettaglio.itertuples(index=False), start=start_row):
            for c_idx, value in enumerate(row, start=start_col):
                worksheet_det.cell(row=r_idx, column=c_idx, value=value)
        for r_idx, row in enumerate(df_report_percentuale.itertuples(index=False), start=start_row):
            for c_idx, value in enumerate(row, start=start_col):
                worksheet_per.cell(row=r_idx, column=c_idx, value=value)


        workbook.save(download_filename)


                # Return the download link using dcc.send_file
        return dcc.send_file(download_filename,filename=download_filename)
                #return dcc.send_file

    return None




@app.callback(
    Output('table-container-2', 'children'),
    [
        Input('dropdown-menu', 'value'),
        Input('start-date-store', 'data'),
        Input('end-date-store', 'data'),
        Input('store-df3', 'data')
    ]
)
def update_table2(selected_region, date_start, date_end,direttore_ore_ambulatorio__dict):
    # Example DataFrame 
    
    df_ore_ambulatorio = pd.DataFrame.from_dict(direttore_ore_ambulatorio__dict)
    for col in df_ore_ambulatorio.select_dtypes(include=['int64', 'float64']).columns:
        df_ore_ambulatorio[col] = df_ore_ambulatorio[col].apply(lambda x: "-" if pd.isna(x) else format_thousands(x))
    num_rows = len(df_ore_ambulatorio)
    
    # Set row height and header height
    row_height = 60  # Estimated row height
    
    # Set max table height for 8 rows plus header
    max_table_height = 6 * row_height + 30
    
    # Determine the table height
    table_height = min(num_rows * row_height + 30, max_table_height)
    
    # Set table style with dynamic height and scroll bar if necessary
    table_style = {
        'height': f'{table_height}px', 
        'overflowY': 'auto' if num_rows > 8 else 'hidden', 
        'width': '100%', 
        'margin': 'auto'
    }
    
    # Define columns
    columns = [{'headerName': col, 'field': col} for col in df_ore_ambulatorio.columns]
    
    return [
        dag.AgGrid(
            id='table',
            columnDefs=columns,
            rowData=df_ore_ambulatorio.to_dict('records'),
            className="ag-theme-alpine",
            rowStyle={'background-color': 'rgba(181, 208, 255, 0.2)'},
            dashGridOptions={
                "rowSelection": "multiple", 
                "suppressRowClickSelection": True, 
                "animateRows": False,
                "defaultColDef": {
                    "resizable": True,
                    "sortable": True,
                    "filter": True,
                    "flex": 1,
                    "rowHeight": row_height
                },
                "scrollContainer": "body",
                "alwaysShowVerticalScroll": True,
                "alwaysShowHorizontalScroll": True,
                "getRowStyle": {
                    'backgroundColor':'#ECF2FE'
                },
            },
            style=table_style
        )
    ]


### TABELLA DIRETTORE - TEMPI MEDI DI ATTESA PER RECLUSTERING
@app.callback(
    Output('table-container-1', 'children'),
    [Input('dropdown-menu', 'value'),
     Input('store-df2','data')
     ]
)
def update_table1(selected_region,direttore_reclustering_tempi_attesa_dict):
    df_recluster_direttore_tempi = pd.DataFrame.from_dict(direttore_reclustering_tempi_attesa_dict)  # Use actual data source
    for col in df_recluster_direttore_tempi.select_dtypes(include=['int64', 'float64']).columns:
        df_recluster_direttore_tempi[col] = df_recluster_direttore_tempi[col].apply(lambda x: "-" if pd.isna(x) else format_thousands(x))

    
    #La colonna si chiama ambito per fare il filtro

# Stampiamo il DataFrame filtrato
    
    num_rows = len(df_recluster_direttore_tempi)
    row_height = 60
    max_table_height =  6 * row_height + 30
    table_height = min(num_rows * row_height + 30, max_table_height)
    table_style = {'height': f'{table_height}px', 'overflowY': 'auto', 'width': '100%', 'margin': 'auto'} if num_rows > 8 else {'height': f'{table_height}px', 'width': '100%', 'margin': 'auto'}
    columns = [{'headerName': col, 'field': col} for col in df_recluster_direttore_tempi.columns]
    
    return dag.AgGrid(
        id='table',
        columnDefs=columns,
        rowData=df_recluster_direttore_tempi.to_dict('records'),
        className="ag-theme-alpine",
        dashGridOptions={
            'suppressRowTransform': True,
            "rowSelection": "multiple",
            "suppressRowClickSelection": True,
            "animateRows": False,
            "defaultColDef": {
                "resizable": True,
                "sortable": True,
                "filter": True,
                "flex": 1,
                "rowHeight": 60
            },
            "alwaysShowVerticalScroll": True,
            "getRowStyle": {
                    'backgroundColor':'#ECF2FE'
                },
        },
        style=table_style,
    )




#### PROVA CALLBACK ALTERNA I DIVERSI LOGIN 

@app.callback(
    Output('output-div_login', 'children'),
    [Input('button_login', 'n_clicks')]
)
def update_output_div(n_clicks):
    global value
    if n_clicks is None:
        # Imposta il valore della variabile su None se il pulsante non è stato ancora cliccato
        value = None
    else:
        # Imposta il valore della variabile su qualcosa di diverso da None quando il pulsante è stato cliccato
        value = 'La variabile è stata valorizzata!'



# Callback per memorizzare l'username associato al pulsante 1
@app.callback(
    Output('username1-store', 'data'),
    [Input('button-1', 'n_clicks')],
    [State('username1', 'value')]
)
def store_username1(button1_clicks, username1):
    if button1_clicks and username1:
        return username1
    else:
        return None


# Callback per memorizzare l'username associato al pulsante 2
@app.callback(
    Output('username2-store', 'data'),
    [Input('button-2', 'n_clicks')],
    [State('username2', 'value')]
)

def store_username2(button2_clicks, username2):
    if button2_clicks and username2:
        return username2
    else:
        return None
# Callback per recuperare l'username memorizzato nei pulsanti 1 e 2 negli store
@app.callback(
    Output('button-1-clicks-store', 'data'),
    [Input('button-1', 'n_clicks')]
)
def store_button_1_clicks(button_1_clicks):
    return button_1_clicks
@app.callback(
    Output('button-2-clicks-store', 'data'),
    [Input('button-2', 'n_clicks')]
)
def store_button_2_clicks(button_2_clicks):
    return button_2_clicks



@app.callback(
    Output('icona', 'children'),
    [Input('button-2-clicks-store', 'data')]
)
def display_icon(button_2_clicks):
    if button_2_clicks and button_2_clicks > 0:
        return html.Div([
            dcc.Link(html.Img(src='/assets/research.png',height = '35px'),    href='/consigli',
                        style={'float': 'right', 'display': 'inline-block','display': 'grid',
    'grid-template-columns': '1fr auto'}
                    )
        ])
    else:
        return None

# Callback per aggiornare il testo data prima disponibilità
@app.callback(
    Output('stima-tempi-attesa', 'children'),
    [Input('dropdown-value-store', 'data'), Input('store-df5','data'),Input('start-date-store','data')]
)
def update_text(codice_agenda,df_data,start_date):
    
    df_prima_disp = pd.DataFrame(df_data) if df_data else pd.DataFrame()
    df_prima_disp['Codice Agenda'] = df_prima_disp['Codice Agenda'].astype(str)
    codice_agenda = str(codice_agenda)
    # Filtrare il dataframe
    filtered_df = df_prima_disp[df_prima_disp['Codice Agenda'] == codice_agenda]

    # Ottenere la prima data dal dataframe filtrato
    if not filtered_df.empty:
        data_riconfigurazione = filtered_df['Data_Prima_disponibilita'].iloc[0]
    else:
        data_riconfigurazione = 'N/A'  # Nel caso in cui non ci siano dati corrispondenti
    start = pd.to_datetime(start_date)
    print(type(start))
    print(type(data_riconfigurazione))
    dr = pd.to_datetime(data_riconfigurazione)
    print(dr)
    # Calcola il mese successivo
    mese_successivo = dr.month + 1
    anno_successivo = dr.year

    # Gestisci il cambio di anno se necessario
    if mese_successivo > 12:
        mese_successivo = 1
        anno_successivo += 1

    # Nome del mese in italiano
    mesi_italiani = [
        'gennaio', 'febbraio', 'marzo', 'aprile', 'maggio', 'giugno',
        'luglio', 'agosto', 'settembre', 'ottobre', 'novembre', 'dicembre'
    ]

    # Ottieni il nome del mese successivo
    nome_mese_successivo = mesi_italiani[mese_successivo - 1]
    frase = f"Gli effetti della riconfigurazione partiranno da "

    return html.Div([
        frase,
        html.B(f"{nome_mese_successivo.capitalize()} {anno_successivo}")
    ])


    #return f"Gli effetti della riconfigurazione partiranno da {nome_mese_successivo.capitalize()} {anno_successivo}"
# Callback per aggiornare l'intervallo di date selezionato
@app.callback(
    Output('start-date-store', 'data'),
    Input('start-day-dropdown', 'value'),
    Input('start-month-dropdown', 'value'),
    Input('start-year-dropdown', 'value')
)
def update_start_date_store(day, month, year):
    formatted_date = f"{year}-{month:02d}-{day:02d}"
    return formatted_date

@app.callback(
    Output('end-date-store', 'data'),
    Input('end-day-dropdown', 'value'),
    Input('end-month-dropdown', 'value'),
    Input('end-year-dropdown', 'value')
)
def update_end_date_store(day, month, year):
    formatted_date = f"{year}-{month:02d}-{day:02d}"
    return formatted_date


@app.callback(
    Output('date-error-message', 'children'),
    [
        Input('start-day-dropdown', 'value'),
        Input('start-month-dropdown', 'value'),
        Input('start-year-dropdown', 'value'),
        Input('end-day-dropdown', 'value'),
        Input('end-month-dropdown', 'value'),
        Input('end-year-dropdown', 'value')
    ]
)
def update_dates(start_day, start_month, start_year, end_day, end_month, end_year):
    
    
    date_start = date(int(start_year), int(start_month), int(start_day))
    date_end = date(int(end_year), int(end_month), int(end_day))
    
    # Check if date_start is greater than date_end
    if date_start > date_end:
        error_message = 'La data di inizio non può essere maggiore della data di fine.'
        return html.Div(error_message, style={'color': 'red'})
    
    
    return ''
"""giorno_odierno = int(oggi.strftime('%d'))
mese_odierno = int(oggi.strftime('%m'))
anno_odierno = int(oggi.strftime('%Y'))"""
data_di_riferimento = date(2024, 10, 27)
"""@app.callback(
    [
        Output('start-day-dropdown', 'options'),
        Output('end-day-dropdown', 'options')
    ],
    [
        Input('start-month-dropdown', 'value'),
        Input('start-year-dropdown', 'value'),
        Input('end-month-dropdown', 'value'),
        Input('end-year-dropdown', 'value')
    ]
)
def update_day_options(start_month, start_year, end_month, end_year):
    # Check if the month and year are provided
    if None in [start_month, start_year, end_month, end_year]:
        raise PreventUpdate
    
    # Filtra le opzioni di giorno per il mese di inizio

    start_days_in_month = calendar.monthrange(start_year, start_month)[1]
    
    start_day_options = [{'label': str(i), 'value': i} for i in range(1, start_days_in_month + 1)]

    if end_year == data_di_riferimento.year and end_month == data_di_riferimento.month:
        end_days_in_month = min(calendar.monthrange(end_year, end_month)[1], data_di_riferimento.day)
    else:
        end_days_in_month = calendar.monthrange(end_year, end_month)[1]

    end_day_options = [{'label': str(i), 'value': i} for i in range(1, end_days_in_month + 1)]

    return start_day_options, end_day_options"""

@app.callback(
    [
        Output('start-month-dropdown', 'options'),
        Output('end-month-dropdown', 'options')
    ],
    [
        Input('start-year-dropdown', 'value'),
        Input('end-year-dropdown', 'value')
    ]
)
def update_date_options(start_year, end_year):
    # Check if the year is provided
    if None in [start_year, end_year]:
        raise PreventUpdate

    # Filtra le opzioni del mese per l'anno di inizio
    if start_year == 2024:
        start_month_options = [{'label': str(i), 'value': i} for i in range(1, data_di_riferimento.month + 1)]
    else:
        start_month_options = [{'label': str(i), 'value': i} for i in range(1, 13)]
    
    # Filtra le opzioni del mese per l'anno di fine
    if end_year == data_di_riferimento.year:
        end_month_options = [{'label': str(i), 'value': i} for i in range(1, data_di_riferimento.month + 1)]
    else:
        end_month_options = [{'label': str(i), 'value': i} for i in range(1, 13)]


    return start_month_options, end_month_options

#### CALCOLO GRADO DI URGENZA 

def data_agende_home(prima_disp_file_Manuel,start_date,end_date):
    ## Funzione per la creazione--> Grado di urgenza** e  file data_agende pagina layoutagende
    
    ### GRADO DI GRADO_SATURAZIONE
    calendar_agende['GRADO_SATURAZIONE'] = calendar_agende['BUSY_SLOTS'] / calendar_agende['MAX_QUANTITY']
    calendar_agende['CALENDAR_DATE'] = pd.to_datetime(calendar_agende['CALENDAR_DATE'])
    calendar_agende_filt = calendar_agende[(calendar_agende['CALENDAR_DATE'] >= start_date) & (calendar_agende['CALENDAR_DATE'] <= end_date)]
    # Raggruppiamo per DIARY_ID e calcoliamo la media di GRADO_SATURAZIONE
    grado_saturazione_avg = calendar_agende_filt.groupby('DIARY_ID')['GRADO_SATURAZIONE'].mean().reset_index()
    
    
    #agende_file = common_path + '\V_AGENDE_ATTIVE_CE.csv'
    grado_saturazione_avg['DIARY_ID'] = grado_saturazione_avg['DIARY_ID'].astype('str')
    ## GRADO DI PRIORITIZZAZIONE
    #df_decodifica= pd.read_csv(path_decodifica_ambiti_garanzia, sep=";", dtype=dtype)
    #df_decodifica.rename(columns={'sts11': 'STS11', 'ambiti_aziende': 'AMBITI_AZIENDE'}, inplace=True)

    timeband_orders_df['START_DATE'] = pd.to_datetime(timeband_orders_df['START_DATE'])

    # Filtriamo il DataFrame per includere solo le righea con date nel range specificato data_agednda start <= data_filtro
    #                                                                                    data_agednda end >= data_filtro
    
    timeband_orders_df_filt =  timeband_orders_df[(timeband_orders_df['START_DATE'] <= end_date) & (timeband_orders_df['END_DATE'] >= start_date)]
    ## da modificare
    
    timeband_orders_df_filt.PRIORITY_CLASS_MASK.fillna(value='BDUP', inplace=True)
    # Funzione per calcolare la percentuale di classi di priorità
    def calculate_priority_percentage(classi_priorita):
        if pd.isnull(classi_priorita):
            return 0
        # Conta il numero di lettere nella stringa delle classi di priorità
        num_letters = len(classi_priorita)
        # Se ci sono più di una lettera, restituisci 0%, altrimenti 100%
        return 0 if num_letters > 1 else 1

    # Applica la funzione alla colonna CLASSI_PRIORITA per ottenere grado_di_prioritizzazione
    # Raggruppa per DIARY_ID e calcola la media di grado_di_prioritizzazione
    timeband_orders_df_filt_no_dup = timeband_orders_df_filt[['TIMEBAND_ID','DIARY_ID','PRIORITY_CLASS_MASK']].drop_duplicates(keep='first')
    timeband_orders_df_filt_no_dup['grado_di_prioritizzazione'] = timeband_orders_df_filt_no_dup['PRIORITY_CLASS_MASK'].apply(calculate_priority_percentage)
    grado_prioritizzazione_avg = timeband_orders_df_filt_no_dup.groupby('DIARY_ID')['grado_di_prioritizzazione'].mean().reset_index()
    grado_prioritizzazione_avg['DIARY_ID'] = grado_prioritizzazione_avg['DIARY_ID'].astype('str')

    ## GRADO DI PRIMA DISPONIBILITA'
    # Raggruppa per DIARY_ID e calcola la media di grado_di_prioritizzazione
    
    
    prima_disp_file_Manuel_rename =prima_disp_file_Manuel.rename(columns={'Codice Agenda': 'DIARY_ID'})
    
    # Converti DIARY_ID in calendar in int64
    calendar_agende['DIARY_ID'] = calendar_agende['DIARY_ID'].astype('str')
    prima_disp_file_Manuel_rename['DIARY_ID'] = prima_disp_file_Manuel_rename['DIARY_ID'].astype('str')
    # Esegui il merge
    joined_df = calendar_agende.merge(prima_disp_file_Manuel_rename,on='DIARY_ID',how='left')
    joined_df['Data_Prima_disponibilita']= pd.to_datetime(joined_df['Data_Prima_disponibilita'],format="%d-%m-%Y")
    joined_df_no_timeband_dup =joined_df.groupby(['CALENDAR_DATE','DIARY_ID','Data_Prima_disponibilita']).first().reset_index()
    numeratore = joined_df_no_timeband_dup[joined_df_no_timeband_dup['CALENDAR_DATE'] <= joined_df_no_timeband_dup['Data_Prima_disponibilita']].groupby('DIARY_ID').size().reset_index(name='count')
    denominatore = joined_df_no_timeband_dup.groupby('DIARY_ID')['CALENDAR_DATE'].size().reset_index(name='denominatore')
    new_df = numeratore.merge(denominatore, on='DIARY_ID',how='left')
    new_df['GRADO DI PRIMA DISP']=((new_df['count']/new_df['denominatore']))
    new_df['DIARY_ID'] = new_df['DIARY_ID'].astype('str')
    ### FAI COMANDARE LA TABELLA DI TUTTE LE AGENDE E TUTTE LE DESCRIZIONI 
    agende_descrizione_join = agende_descrizione.copy()
    agende_descrizione_join['Codice Agenda'] = agende_descrizione_join['Codice Agenda'].astype('str')
    agende_descrizione_join.rename(columns ={'Codice Agenda' : 'DIARY_ID'},inplace=True)
    join_agenda_grado_prima_disp = agende_descrizione_join.merge(new_df,on='DIARY_ID',how='left')
    join_prima_disp_saturazione = join_agenda_grado_prima_disp.merge(grado_saturazione_avg,on='DIARY_ID',how='left')
    join_prima_disp_saturaz_prioritizz= join_prima_disp_saturazione.merge(grado_prioritizzazione_avg,on='DIARY_ID',how='left')
    join_prima_disp_saturaz_prioritizz.grado_di_prioritizzazione.fillna(value=0, inplace=True)
    ### CALCOLO GRADO DI URGENZA**
    def calculate_value(row):
        if pd.isnull(row['GRADO DI PRIMA DISP']) or pd.isnull(row['GRADO_SATURAZIONE'])or pd.isnull(row['grado_di_prioritizzazione']):
            return np.nan
        else:
            return round(1 -(0.2 * row['GRADO DI PRIMA DISP'] + 0.4 * row['GRADO_SATURAZIONE'] + 0.4 * row['grado_di_prioritizzazione'] ),2)
    join_prima_disp_saturaz_prioritizz['Grado di urgenza**'] = join_prima_disp_saturaz_prioritizz.apply(calculate_value, axis=1)
    grado_di_urgenza = join_prima_disp_saturaz_prioritizz[['DIARY_ID','Grado di urgenza**']]
    grado_di_urgenza.rename(columns = {'DIARY_ID':'Codice Agenda'},inplace=True)
    
    ### CALCOLO DATA_AGENDE UNENDO I DF CREATI (COD/DESC AGENDA - GRADO URGENZA - ADERENZA GRAFO - SENTINELLA)
    
    ###CALCOLO COLONNA SENTINELLA 
    sentinella_2 = sentinella.rename(columns = {'Codice Catalogo regionale' : 'Codice Prestazione'})
    sentinella_2['flag'] = 'Si'
    agende_copy = prestazioni_agende_df.copy()
    sentinella_2['Codice Prestazione'] = sentinella_2['Codice Prestazione'].astype('str')
    agende_copy['Codice Prestazione'] = agende_copy['Codice Prestazione'].astype('str')
    agende_con_sentinella = agende_copy.merge(sentinella_2, on = 'Codice Prestazione', how ='left')
    df_sentinella = agende_con_sentinella.groupby('Codice Agenda').apply(lambda x: 'Si' if (x['flag'] == 'Si').any() else 'No').reset_index(name ='Sentinella')
    ## AGGIUNTA GRADO DI ADERENZA AL GRAFO **
    df_chart_1 = df_chart.rename(columns = {'Value' : 'Grado di aderenza al grafo*'})
    ## Parti dal file agende_descrizione che contiene tutte le agende con la descrizioni e fai tutte left
    agende_descrizione['Codice Agenda'] = agende_descrizione['Codice Agenda'].astype('str')
    df_chart_1['Codice Agenda'] = df_chart_1['Codice Agenda'].astype('str')
    codice_descrizione = agende_descrizione.merge(df_chart_1, on ='Codice Agenda', how='left')
    urg_aderenza = codice_descrizione.merge(grado_di_urgenza, on ='Codice Agenda', how='left')
    df_data_agende = urg_aderenza.merge(df_sentinella, on ='Codice Agenda', how='left')
    """df_data_agende['Grado di aderenza al grafo*'] *= 100
    df_data_agende['Grado di urgenza**'] *= 100
    df_data_agende['Grado di aderenza al grafo*'] = df_data_agende['Grado di aderenza al grafo*'].apply(lambda x: '-' if pd.isna(x) else f"{x:.0f}%")
    df_data_agende['Grado di urgenza**'] = df_data_agende['Grado di urgenza**'].apply(lambda x: '-' if pd.isna(x) else f"{x:.0f}%")"""
    ## EXPORT DEL DF data_agende per poterlo manipolare
    #df_data_agende.to_excel(os.path.join(root_path, 'data_agende.xlsx'))
    
    ## Aggiunta ambito 
    
    print('---FINEE FUNZIONE---')
    
    return df_data_agende

def data_prima_disponibilita(df_prima_disponibilita):
    ## Funzione per la creazione df prima disponibilità pagina 4 tables righetto in basso
    df_prima_disponibilita['Data_Prima_disponibilita'] = pd.to_datetime(df_prima_disponibilita['Data_Prima_disponibilita'])

    # Convert the datetime column to the desired format
    df_prima_disponibilita['Data_Prima_disponibilita'] = df_prima_disponibilita['Data_Prima_disponibilita'].dt.strftime('%d-%m-%Y')
    df_prima_disponibilita_df = df_prima_disponibilita.reset_index()
    df_prima_disponibilita_df.rename(columns = {'id_agenda':'Codice Agenda'},inplace=True)
    
    return df_prima_disponibilita_df

def direttore_ore_ambulatorio(ore_ambulatorio_da_aggiungere):
    ## Funzione per creazione df ore ambulatorio da aggiungere 
    
    ore_ambulatorio_da_aggiungere['Giorno'] = pd.to_datetime(ore_ambulatorio_da_aggiungere['Giorno'])
    # Aggiungere una colonna 'mese' per facilitare il raggruppamento
    ore_ambulatorio_da_aggiungere['mese'] = ore_ambulatorio_da_aggiungere['Giorno'].dt.to_period('M')
    # Ordinare il DataFrame per 'mese' e 'giorno'
    ore_ambulatorio_da_aggiungere_sorted =  ore_ambulatorio_da_aggiungere.sort_values(by=['mese', 'Re_cluster', 'Giorno'])
    # Prendere l'ultima riga di ogni gruppo 'mese'
    ore_ambulatorio_da_aggiungere_last_day = ore_ambulatorio_da_aggiungere_sorted.groupby(['mese', 'Re_cluster']).tail(1).reset_index(drop=True)
    ore_ambulatorio_da_aggiungere_last_day['Re_cluster'] = ore_ambulatorio_da_aggiungere_last_day['Re_cluster'].astype('str')
    decodifica_descrizione.rename(columns = {'Re-clustering' : 'Re_cluster'},inplace=True)
    ore_ambulatorio_da_aggiungere_group_2 = ore_ambulatorio_da_aggiungere_last_day.merge(decodifica_descrizione, on ='Re_cluster', how='left')
    ore_ambulatorio_da_aggiungere_group_2['ora'] = ore_ambulatorio_da_aggiungere_group_2['minuti_ritardo']/60
    # Dizionario per la conversione dei mesi
    month_map = {
        '01': "Gennaio",
        '02': "Febbraio",
        '03': "Marzo",
        '04': "Aprile",
        '05': "Maggio",
        '06': "Giugno",
        '07': "Luglio",
        '08': "Agosto",
        '09': "Settembre",
        '10': "Ottobre",
        '11': "Novembre",
        '12': "Dicembre"
    }

    ## Funzione per creare la colonna desc_mese
    def convert_month_year(mese):
        year, month = mese.split('-')
        return f"{month_map[month]}'{year[-2:]}"

    # Applicare la funzione a ogni elemento della colonna 'mese'
    ore_ambulatorio_da_aggiungere_group_2['mese_2'] = ore_ambulatorio_da_aggiungere_group_2['mese'].astype('str')
    ore_ambulatorio_da_aggiungere_group_2['desc_mese'] = ore_ambulatorio_da_aggiungere_group_2['mese_2'].apply(convert_month_year)

    ore_ambulatorio_da_aggiungere_group_2['ora'] = ore_ambulatorio_da_aggiungere_group_2['ora'].astype('int32')
    ore_ambulatorio_da_aggiungere_group_2['ora'] = round(ore_ambulatorio_da_aggiungere_group_2['ora'],0)

    ordered_months = [
        "Gennaio'24", "Febbraio'24", "Marzo'24", "Aprile'24", "Maggio'24", "Giugno'24",
        "Luglio'24", "Agosto'24", "Settembre'24", "Ottobre'24", "Novembre'24", "Dicembre'24",
            "Gennaio'25", "Febbraio'25", "Marzo'25", "Aprile'25", "Maggio'25", "Giugno'25",
        "Luglio'25", "Agosto'25", "Settembre'25", "Ottobre'25", "Novembre'25", "Dicembre'25",
    ]

    ## Trasposizione del DataFrame per rendere i mesi colonne
    df_transposed = ore_ambulatorio_da_aggiungere_group_2.pivot(index=['Re_cluster', 'Descrizione post-Reclustering'], columns='desc_mese', values = 'ora')

    ## Resetta l'indice per ottenere un DataFrame classico
    df_transposed.reset_index(inplace=True)

    df_transposed = df_transposed.reindex(columns=['Re_cluster', 'Descrizione post-Reclustering']  + [col for col in ordered_months if col in df_transposed.columns])
    for col in df_transposed.columns[2:]:
        df_transposed[col] = pd.to_numeric(df_transposed[col], errors='coerce').astype('Int64')
        
    df_transposed.rename(columns = {'Descrizione post-Reclustering' : 'Tipo di ambulatorio/​attrezzatura' }, inplace=True)
    data_tabella_ore_di_ambulatorio =  df_transposed.drop('Re_cluster',axis=1)
    data_tabella_ore_di_ambulatorio = data_tabella_ore_di_ambulatorio[(data_tabella_ore_di_ambulatorio['Tipo di ambulatorio/​attrezzatura'] != 'DIALISI') & (data_tabella_ore_di_ambulatorio['Tipo di ambulatorio/​attrezzatura'] != 'PACC')]
    return data_tabella_ore_di_ambulatorio
    


def direttore_tabella_tempi_di_attesa(reclustering_tempi_attesa,df_prima_disponibilita_reclusterizzato,start_date,end_date): #modifiche Antonio
    ## Funzione per creazione tabella tempi di attesa reclustering 
    start_date = pd.to_datetime(start_date)
    end_date = pd.to_datetime(end_date)
    decodifica_descrizione = re_cluster[['Re-clustering','Descrizione post-Reclustering']].drop_duplicates()
    prenot['codice_priorità'].replace(['D1','D2'], ['D','D'],inplace = True)
    prenot_solo_prenotato = prenot[prenot['stato_appuntamento'] == 'Prenotata']
    decodifica_descrizione = re_cluster[['Re-clustering','Descrizione post-Reclustering']].drop_duplicates()
    decodifica_recluster = re_cluster[['Codice Catalogo','Descrizione Catalogo','Re-clustering']]
    decodifica_recluster.rename(columns = {'Codice Catalogo' : 'codice_prestazione_cur'},inplace=True)
    ## Rimuovi duplicati
    decodifica_recluster = decodifica_recluster.drop_duplicates()
    ## Aggiungi informazioni recluster
    prenot_with_recluster = prenot_solo_prenotato.merge(decodifica_recluster[['codice_prestazione_cur','Re-clustering']], on = 'codice_prestazione_cur', how ='left')
    # Calcola differenza giorni 
    prenot_with_recluster['data_diff'] = prenot_with_recluster['d_data_appuntame'] - prenot_with_recluster['d_data_contatto']
    prenot_with_recluster['Re-clustering'] = prenot_with_recluster['Re-clustering'].astype('str')
    ## Ragguppamento media su prio- re_cluster e 
    tab_recluster = prenot_with_recluster.groupby(['ambiti_aziende','Re-clustering','codice_priorità'])['data_diff'].mean().reset_index(name = 'Attuale')
    # convertito in giorni 
    tab_recluster['Attuale'] = tab_recluster['Attuale'].dt.days
    #decodifica_descrizione['Re-clustering'] =decodifica_descrizione['Re-clustering'].astype('str')
    #tab_recluster_fin = tab_recluster.merge(decodifica_descrizione, on = 'Re-clustering', how ='left')
    ## NON TUTTO ENTRA NEL RECLUSTER QUINDI ALCUNI GRUPPI DI PRESTAZIONI NON HANNO RECLUSTER 
    #recluster_attuale = tab_recluster_fin[tab_recluster_fin['Re-clustering'] != 'nan']
    tab_recluster.rename(columns = {'ambiti_aziende': 'ambito','codice_priorità': 'Priorità'},inplace=True)
    attesa_2_new = reclustering_tempi_attesa.rename(columns ={'classe_priorita' : 'Priorità' ,'attesa_media' : 'Post-Riconfigurazione'})
    attesa_2_new['Post-Riconfigurazione'] = attesa_2_new['Post-Riconfigurazione'].astype('int32')

    #MERGE STATO ATTUALE CON POST-RICONFIGURAZIONE 
    total_finale = attesa_2_new[['Re-clustering','Priorità','Post-Riconfigurazione','ambito']].merge(tab_recluster, on =['Re-clustering','Priorità','ambito'], how ='left')
    decodifica_descrizione['Re-clustering'] =decodifica_descrizione['Re-clustering'].astype('str')
    ## ELIMINIAMO I RECLUSTER CHE NON HANNO RICONFIGURAZIONE 
    total_finale_desc = total_finale.merge(decodifica_descrizione, on = 'Re-clustering', how ='left')
    not_null = total_finale_desc[total_finale_desc['Post-Riconfigurazione'].notnull()]
    not_null['Post-Riconfigurazione'] = not_null['Post-Riconfigurazione'].astype('int32')
    b = not_null.rename(columns = {'Descrizione post-Reclustering': 'Gruppi di prestazioni','Priorità' : 'Classi di Priorità'})
    recluster_tempi_attesa_fin = b[['Re-clustering','ambito','Gruppi di prestazioni','Classi di Priorità' ,'Attuale' ,'Post-Riconfigurazione']]
    df_prima_disponibilita_reclusterizzato = df_prima_disponibilita_reclusterizzato.reset_index()
    df_prima_disponibilita_reclusterizzato.rename(columns = {'Re_cluster':'Re-clustering'},inplace=True)
    
    merge_recluster_final = recluster_tempi_attesa_fin.merge(df_prima_disponibilita_reclusterizzato, on =['Re-clustering','ambito'], how ='left')
    def calculate_month_diff(start_date, end_date):
        if pd.isna(end_date):
            return 'Non efficientabile nel periodo selezionato'
        diff_anni =  end_date.year - start_date.year 
        diff_mesi =  end_date.month +1 -start_date.month 
        total_months = diff_anni * 12 + diff_mesi 
        if total_months == 1: 
            return "1 mese"
        else:
            return f"{total_months} mesi"
    merge_recluster_final['Data_Prima_disponibilita'] = pd.to_datetime(merge_recluster_final['Data_Prima_disponibilita'])
    merge_recluster_final['Effetti della riconfigurazione a partire da*'] = merge_recluster_final['Data_Prima_disponibilita'].apply(lambda x: calculate_month_diff(start_date, x))
    merge_recluster_final_2 = merge_recluster_final[['Gruppi di prestazioni',
       'Classi di Priorità', 'Attuale', 'Post-Riconfigurazione',
       'Effetti della riconfigurazione a partire da*']]
    ### FILTRO NO PACC E DIALISI
    merge_recluster_final_2 = merge_recluster_final_2[(merge_recluster_final_2['Gruppi di prestazioni'] != 'PACC') & (merge_recluster_final_2['Gruppi di prestazioni'] != 'DIALISI')]
    return merge_recluster_final_2

def transform_4_tables(calendar_agende,recap_post_riconfigurazione,tempi_attesa_post_riconfigurazione,start_date,end_date):
    ## Funzione per creazione 4 tables 
    
    calendar_agende['CALENDAR_DATE'] = pd.to_datetime(calendar_agende['CALENDAR_DATE'])

    # Filtriamo il DataFrame per includere solo le righe con date nel range specificato
    calendar_agende = calendar_agende[(calendar_agende['CALENDAR_DATE'] >= start_date) & (calendar_agende['CALENDAR_DATE'] <= end_date)]
    #calendar_agende['SLOT_SIZE'] = pd.to_numeric(calendar_agende['SLOT_SIZE'])
    calendar_agende['FREE_SLOTS'] = pd.to_numeric(calendar_agende['FREE_SLOTS'])

    ## Slot Attuale
    calendar_agende['Slot_Attuale'] =  calendar_agende['FREE_SLOTS']

    # Raggruppa il DataFrame per la colonna 'DIARY_ID' e somma solo i valori della colonna 'PRODUCT'
    recap_stato_attuale_elab = calendar_agende.groupby('DIARY_ID')['Slot_Attuale'].sum().reset_index()

    recap_stato_attuale_elab['DIARY_ID'] = recap_stato_attuale_elab['DIARY_ID'].astype('str')
    recap_stato_attuale_elab.rename(columns = {'DIARY_ID' : 'Codice Agenda'},inplace=True)


    ### ******** ATTENZIONA QUESTO PUNTO *******
    #priorita.groupby('DIARY_ID')['CLASSI_PRIORITA'].nunique().reset_index().query('CLASSI_PRIORITA >1')
    decod_prio = priorita.groupby('DIARY_ID')['CLASSI_PRIORITA'].first().reset_index()
    decod_prio.CLASSI_PRIORITA.fillna(value='BDUP', inplace=True)
    decod_prio['DIARY_ID'] = decod_prio['DIARY_ID'].astype('str')
    decod_prio.rename(columns = {'DIARY_ID' : 'Codice Agenda'},inplace=True)

    recap_stato_attuale_elab = recap_stato_attuale_elab.merge(decod_prio,on ='Codice Agenda', how ='left')
    recap_stato_attuale_elab.rename(columns = {'CLASSI_PRIORITA' : 'Priorità'},inplace=True)

    # Funzione per esplodere i caratteri in righe separate mantenendo le altre colonne
    def esplodi_caratteri(df, colonna):
        exploded_df = df[colonna].apply(list).explode()
        repeated_columns = df.drop(columns=[colonna]).loc[df.index.repeat(df[colonna].str.len())].reset_index(drop=True)
        result_df = pd.concat([exploded_df.reset_index(drop=True), repeated_columns], axis=1)
        result_df.columns = [colonna] + list(repeated_columns.columns)  
        
        return result_df

    # Applichiamo la funzione al DataFrame
    recap_stato_attuale = esplodi_caratteri(recap_stato_attuale_elab, 'Priorità')
    recap_stato_attuale.replace(['U','B','D','P'],['Urgente','Breve','Differibile','Programmata'],inplace =True)

    ## RECAP - POST-RICONFIGURAZIONE  ##df_daily_requests_in_minutes_per_agend

    recap_post_riconfigurazione_renamed = recap_post_riconfigurazione.rename(columns = {'Slot_assegnati_U' : 'Urgente' ,'Slot_assegnati_B' : 'Breve', 'Slot_assegnati_D' :'Differibile','Slot_assegnati_P' : 'Programmata','Codice_agenda' : 'Codice Agenda'  })
    asseganzione_melted = recap_post_riconfigurazione_renamed.melt(id_vars=['Giorno', 'ambito', 'Codice Agenda'],
                        value_vars=['Urgente', 'Breve', 'Differibile', 'Programmata'],
                        var_name='Priorità',
                        value_name='Slot_Riconfigurazione')
    asseganzione_melted['Giorno'] = pd.to_datetime(asseganzione_melted['Giorno'])
    recap_post_riconfigurazione_group = asseganzione_melted.groupby(['Codice Agenda','Priorità'])['Slot_Riconfigurazione'].sum().reset_index(name='Slot_Riconfigurazione')
    recap_post_riconfigurazione_group['Codice Agenda'] = recap_post_riconfigurazione_group['Codice Agenda'].astype('str')


    ## TEMPI MEDI STIMATI - ATTUALE   ## da dataframe 
    ## INTEGRA
    #prenot = pd.read_parquet(r'C:\Users\lucamartino\OneDrive - KPMG\So.Re.Sa\Data Ingestion\CUP_Forecast\cup_prenotato_cleaned_wave_1.parquet')
    prenot['codice_priorità'].replace(['D1','D2'], ['D','D'],inplace = True)
    prenot_solo_prenotato = prenot[prenot['stato_appuntamento'].isin(['Prenotata'])]
    prenot_solo_prenotato['Giorni_Attuale'] = (prenot_solo_prenotato['d_data_appuntame'] - prenot_solo_prenotato['d_data_contatto']).dt.days
    tempi_medi_attuale = prenot_solo_prenotato.groupby(['codice_agenda', 'codice_priorità']).agg({
        'Giorni_Attuale': 'mean'  
    }).reset_index()
    tempi_medi_attuale['Giorni_Attuale'] = tempi_medi_attuale['Giorni_Attuale'].round(0).astype(int)
    tempi_medi_attuale.replace(['U','B','D','P'],['Urgente','Breve','Differibile','Programmata'],inplace =True)
    tempi_medi_attuale.rename(columns = {'codice_agenda' : 'Codice Agenda', 'codice_priorità' : 'Priorità'}, inplace=True)


    ## TEMPI MEDI STIMATI - POST-RICONFIGURAZIONE   ##df_attesa_agenda_eapoc_combined
    tempi_attesa_post_riconfigurazione['classe_priorita'].replace(['P','U','B','D'], ['Programmata','Urgente','Breve','Differibile'],inplace=True)
    tempi_attesa_post_riconfigurazione_ren = tempi_attesa_post_riconfigurazione.rename(columns ={'codice_agenda' : 'Codice Agenda','classe_priorita' : 'Priorità' ,'attesa_media' : 'Giorni_Riconfigurazione'})
    tempi_attesa_post_riconfigurazione_ren['Giorni_Riconfigurazione'] = tempi_attesa_post_riconfigurazione_ren['Giorni_Riconfigurazione'].astype('int32')
    tempi_attesa_post_riconfigurazione_ult = tempi_attesa_post_riconfigurazione_ren[['Codice Agenda','Priorità','Giorni_Riconfigurazione']]

    ## JOIN FINALE PER OUTPUT 
    recap_tot = recap_stato_attuale.merge(recap_post_riconfigurazione_group, on =['Codice Agenda','Priorità'], how ='outer')
    tempi_tot = tempi_medi_attuale.merge(tempi_attesa_post_riconfigurazione_ult, on =['Codice Agenda','Priorità'], how ='outer')
    all_4_tables = recap_tot.merge(tempi_tot, on =['Codice Agenda','Priorità'], how ='outer')

    return all_4_tables




@app.callback(
    [Output('store-df1', 'data'),
     Output('store-df2', 'data'),
     Output('store-df3', 'data'),
     Output('store-df4', 'data'),
     Output('store-df5', 'data'),
     Output('store-df6', 'data'),
     Output('url-redirect', 'pathname')],
    [Input('button-avanti', 'n_clicks'),
    Input('start-day-dropdown', 'value'),
    Input('start-month-dropdown', 'value'),
    Input('start-year-dropdown', 'value'),
    Input('end-day-dropdown', 'value'),
    Input('end-month-dropdown', 'value'),
    Input('end-year-dropdown', 'value'),
    Input('button-2-clicks-store', 'data'),
    Input('button-1-clicks-store', 'data'),
    Input('initial-dropdown', 'value'), ### ambito per filtro df_agende legato al file caricato
    Input('upload-data', 'contents'),   ## codice_agenda per filtro df_agende legato al file_caricato
    State('upload-data', 'filename')    ## codice_agenda per filtro df_agende legato al file_caricato
    
    ]
    )



def generate_data_and_redirect(n_clicks,startday,startmonth,startyear,endday,endmonth,endyear,button_2_clicks,button_1_clicks,ambito_filtro,contents,filename):
    
    if contents is not None:
        df_filtro_file = parse_contents(contents, filename)
        print('Filtro file', df_filtro_file)
        if df_filtro_file is not None and 'CODICE_AGENDA' in df_filtro_file.columns:
            unique_values = df_filtro_file['CODICE_AGENDA'].unique()
            unique_values = unique_values.astype(str)
            print('Valori unici da utilizzare ??', unique_values)
    if n_clicks is None:
        raise PreventUpdate
    datainiziale = str(startyear) + '-' + str(startmonth) + '-' + str(startday)
    datainiziale_dt = pd.to_datetime(datainiziale).date()
    datainiziale = str(datainiziale_dt)
    datafinale = str(endyear) + '-' + str(endmonth) + '-' + str(endday)
    datafinale_dt = pd.to_datetime(datafinale).date()
    datafinale = str(datafinale_dt)
    print(datainiziale)
    print(datafinale)
    def main_new(ambito,data_iniziale,data_finale,lista_agende,df_agende_unique_fe): 
            start_day = '2024-01-01'         #### NOTE_PER_AGGIORNAMENTO -> questa data dovra essere allineata con il today dello smart scheduler
            if ambito == 'ASL Caserta A01': 
                scheduler = scheduler_a01
            elif ambito == 'ASL Caserta A02':
                scheduler = scheduler_a02
            elif ambito == 'ASL Caserta A03':
                scheduler = scheduler_a03
            elif ambito == 'ASL Caserta A04':
                scheduler = scheduler_a04
        #    #strutture_private_accreditate_uniche = get_strutture_private_accreditate_uniche(input_path=SCHEDULER_ABSOLUTE_PATH+r"/INPUT/", output_path=SCHEDULER_ABSOLUTE_PATH+f"/OUTPUT/{ambito}")
        #    #df_cluster_descrizione = get_durata_slot_assegnato_cluster(dict_duration_cluster, df_cluster, input_path=SCHEDULER_ABSOLUTE_PATH+r"/INPUT/", output_path=SCHEDULER_ABSOLUTE_PATH+f"/OUTPUT/{ambito}")
        #    df_prima_disponibilita = create_dataframe_prime_date_disponibili_agende(scheduler.agenda_manager.agende_senza_fascie, data_iniziale, output_path=SCHEDULER_ABSOLUTE_PATH+f"/OUTPUT/{ambito}")
        #    prima_data_disponibile_int = convert_df_prima_data_disponibile_a_int(df_prima_disponibilita)
        #    all_processes_filtered = filter_all_processes(scheduler.queue_manager.all_processes, start_day, data_iniziale, data_finale)
        #    agende_associate_set = get_agende_utilizzate(all_processes=scheduler.queue_manager.all_processes)
        #    #file_c_df = read_file_c(cartella_file_sorgente = SCHEDULER_ABSOLUTE_PATH+r"/INPUT/")
        #    df_cup_eapoc_filtered_date = filter_df_date_init_fine(df_cup_eapoc, 'd_data_appuntame', data_iniziale, data_finale)
        #    df_prima_disponibilita_recluster = compute_first_available_date_agende_reclusterizzate(df_cup_eapoc_filtered_date, df_prima_disponibilita, ambito, flag_to_save=False, output_path=SCHEDULER_ABSOLUTE_PATH+f"/OUTPUT/{ambito}")
        #    df_cup_eapoc_filtered_agenda = filter_and_compute_attesa_eapoc(df_cup_eapoc_filtered_date.copy(deep=True), agende_associate_set, df_prima_disponibilita, data_iniziale, 'codice_agenda')
        #    df_cup_eapoc_filtered_recluster = filter_and_compute_attesa_eapoc(df_cup_eapoc_filtered_date.copy(deep=True), agende_associate_set, df_prima_disponibilita_recluster, data_iniziale, 'Re-clustering')
        #    #df_attesa_media_recluster_eaprec = process_cup_ex_ante_pre_configurazione(df_reclustering, agende_associate_set, ambito, output_path=SCHEDULER_ABSOLUTE_PATH+f"/OUTPUT/{ambito}", input_path=SCHEDULER_ABSOLUTE_PATH+r"/INPUT/")
        #    #df_attesa_media_recluster_ex_post_preconfigurazione = process_cup_erogato_ex_post_preconfigurazione(agende_associate_set, df_reclustering, ambito, output_path=SCHEDULER_ABSOLUTE_PATH+f"/OUTPUT/{ambito}", input_path=SCHEDULER_ABSOLUTE_PATH+r"/INPUT/")
        #    # eapoc: ex-ante post-reconfigurazione. eaprec: ex-ante pre-reconfigurazione. epprec: ex-post pre-reconfigurazione. 
        #    df_attesa_reclustering_eapoc_combined = process_tempi_medi_stimati_recluster_eapoc(all_processes_filtered, df_cup_eapoc_filtered_recluster, df_reclustering, df_prima_disponibilita_recluster, ambito, start_day, flag_to_save=True, output_path=SCHEDULER_ABSOLUTE_PATH+f"/OUTPUT/{ambito}")
        #    df_attesa_agenda_eapoc_combined = process_tempi_medi_stimati_agenda_attesa_post_riconfigurazione(all_processes_filtered, df_cup_eapoc_filtered_agenda, ambito, flag_to_save=False, output_path=SCHEDULER_ABSOLUTE_PATH+f"/OUTPUT/{ambito}")
        #    df_daily_requests_in_minutes_per_agend = process_assegnazioni_giornaliere_per_agenda(scheduler.agenda_manager.agende_senza_fascie, calendar_fe, start_day, ambito, data_iniziale, data_finale, flag_to_save=False, output_path=SCHEDULER_ABSOLUTE_PATH+f"/OUTPUT/{ambito}")
        #    df_reclustering_ritardi = process_ore_ambulatoriali(scheduler.queue_manager.storico_queue, df_cluster, df_reclustering, start_day, ambito, data_iniziale, data_finale, flag_to_save=False, output_path=SCHEDULER_ABSOLUTE_PATH+f"/OUTPUT/{ambito}")
        #    return df_attesa_reclustering_eapoc_combined, df_attesa_agenda_eapoc_combined, df_daily_requests_in_minutes_per_agend, df_reclustering_ritardi, df_prima_disponibilita,df_prima_disponibilita_recluster 
        


            print('debug', df_agende_unique_fe)
            dict_duration_cluster = filter_clusters(df_agende_unique_fe, df_cluster, ambito, lambda x: x.mode().values[0])

            strutture_private_accreditate_uniche = get_strutture_private_accreditate_uniche(input_path=SCHEDULER_ABSOLUTE_PATH+r"/INPUT/", output_path=SCHEDULER_ABSOLUTE_PATH+f"/OUTPUT/{ambito}")
            df_cluster_descrizione = get_durata_slot_assegnato_cluster(dict_duration_cluster, df_cluster, input_path=SCHEDULER_ABSOLUTE_PATH+r"/INPUT/", output_path=SCHEDULER_ABSOLUTE_PATH+f"/OUTPUT/{ambito}")
            all_processes_filtered = filter_all_processes(scheduler.queue_manager.all_processes, start_day, data_iniziale, data_finale)
            agende_output_filtered = filter_agende_output(scheduler.agenda_manager.agende_senza_fascie, lista_agende)
            try:
                assert len(agende_output_filtered) > 0
            except:
                print("Selezionare agende da filtrare valide, controllare l'ambito.")
                agende_output_filtered = scheduler.agenda_manager.agende_senza_fascie

            agende_associate_set = get_agende_utilizzate(all_processes=scheduler.queue_manager.all_processes)
            df_prima_disponibilita = create_dataframe_prime_date_disponibili_agende(agende_output_filtered, data_iniziale, output_path=SCHEDULER_ABSOLUTE_PATH+f"/OUTPUT/{ambito}")
            prima_data_disponibile_int = convert_df_prima_data_disponibile_a_int(df_prima_disponibilita)
            #df_reclustering = load_reclustering_data(file_path=SCHEDULER_ABSOLUTE_PATH+r"/INPUT/", file_name="Re-clustering_Spec_Fa.Re_v.2.0_20240624.xlsx", sheet_name='Spectral_clustering_600_limited')
            df_agende_unique_fe = add_recluster_column_to_df(df_agende_unique_fe, df_reclustering, cod_prestazione_nome_col='SERVICE_REGIONAL_CODE')
            df_prima_disponibilita_recluster = compute_first_available_date_agende_reclusterizzate(df_agende_unique_fe, df_prima_disponibilita, ambito, flag_to_save=False, output_path=SCHEDULER_ABSOLUTE_PATH+f"/OUTPUT/{ambito}")
            df_cup_eapoc_filtered_date = filter_df_date_init_fine(df_cup_eapoc, 'd_data_appuntame', data_iniziale, data_finale)
            df_cup_eapoc_filtered_agenda = filter_and_compute_attesa_eapoc(df_cup_eapoc_filtered_date.copy(deep=True), agende_associate_set, df_prima_disponibilita, data_iniziale, 'codice_agenda')
            df_cup_eapoc_filtered_recluster = filter_and_compute_attesa_eapoc(df_cup_eapoc_filtered_date.copy(deep=True), agende_associate_set, df_prima_disponibilita_recluster, data_iniziale, 'Re-clustering')

            #df_attesa_media_recluster_eaprec = process_cup_ex_ante_pre_configurazione(df_reclustering, agende_associate_set, ambito, output_path=SCHEDULER_ABSOLUTE_PATH+f"/OUTPUT/{ambito}", input_path=SCHEDULER_ABSOLUTE_PATH+r"/INPUT/")
            #df_attesa_media_recluster_ex_post_preconfigurazione = process_cup_erogato_ex_post_preconfigurazione(agende_associate_set, df_reclustering, ambito, output_path=SCHEDULER_ABSOLUTE_PATH+f"/OUTPUT/{ambito}", input_path=SCHEDULER_ABSOLUTE_PATH+r"/INPUT/")
            # eapoc: ex-ante post-reconfigurazione. eaprec: ex-ante pre-reconfigurazione. epprec: ex-post pre-reconfigurazione. 
            df_attesa_reclustering_eapoc_combined = process_tempi_medi_stimati_recluster_eapoc(all_processes_filtered, df_cup_eapoc_filtered_recluster, df_reclustering, df_prima_disponibilita_recluster, ambito, start_day, flag_to_save=False, output_path=SCHEDULER_ABSOLUTE_PATH+f"/OUTPUT/{ambito}")
            df_attesa_agenda_eapoc_combined = process_tempi_medi_stimati_agenda_attesa_post_riconfigurazione(all_processes_filtered, df_cup_eapoc_filtered_agenda, ambito, flag_to_save=False, output_path=SCHEDULER_ABSOLUTE_PATH+f"/OUTPUT/{ambito}")
            df_daily_requests_in_minutes_per_agend = process_assegnazioni_giornaliere_per_agenda(agende_output_filtered, calendar_fe, start_day, ambito, data_iniziale, data_finale, flag_to_save=False, output_path=SCHEDULER_ABSOLUTE_PATH+f"/OUTPUT/{ambito}")
            df_reclustering_ritardi = process_ore_ambulatoriali(scheduler.queue_manager.storico_queue, df_cluster, df_reclustering, start_day, ambito, data_iniziale, data_finale, flag_to_save=False, output_path=SCHEDULER_ABSOLUTE_PATH+f"/OUTPUT/{ambito}")

            return (df_attesa_reclustering_eapoc_combined,
                    df_attesa_agenda_eapoc_combined,
                    df_daily_requests_in_minutes_per_agend,
                    df_reclustering_ritardi,
                    df_prima_disponibilita,
                    df_prima_disponibilita_recluster)
                
        
        
        ## LANCIO FUNZIONE
    (df_attesa_reclustering_eapoc_combined, ##direttore tabella 1
        df_attesa_agenda_eapoc_combined,       ## 4 tables -post riconfigurazione tempi attesa
        df_daily_requests_in_minutes_per_agend,## 4tables -post riconfigurazione slot 
        df_reclustering_ritardi,              ## direttore ore di ambulatorio
        df_prima_disponibilita,               ## prima disp
        df_prima_disponibilita_reclusterizzato)  = main_new(ambito=ambito_filtro,
            data_iniziale = datainiziale,
            data_finale = datafinale,
            lista_agende = unique_values,
            df_agende_unique_fe=df_agende_unique_fe)
        #print(df_prima_disponibilita_reclusterizzato)
        # ## SOLO PER DIRETTORE
        # ## FILTRO
 ## OPERATORE

    df_attesa_agenda_eapoc_combined=df_attesa_agenda_eapoc_combined[df_attesa_agenda_eapoc_combined['solo_cup'] != True]
        
        
    four_tables = transform_4_tables(calendar_agende,df_daily_requests_in_minutes_per_agend,df_attesa_agenda_eapoc_combined,start_date=datainiziale,end_date=datafinale)
    df_prima_disponibilita_df = data_prima_disponibilita(df_prima_disponibilita)
    df_data_agende = data_agende_home(df_prima_disponibilita_df,start_date=datainiziale,end_date=datafinale)
    decodifica_agende_ambito['Codice Agenda'] = decodifica_agende_ambito['Codice Agenda'].astype('str')
    df_data_agende['Codice Agenda'] = df_data_agende['Codice Agenda'].astype('str')
    df_data_agende_con_ambito = df_data_agende.merge(decodifica_agende_ambito, on = 'Codice Agenda', how='left')
    df_data_agende_filtered = df_data_agende_con_ambito[(df_data_agende_con_ambito['AMBITI_AZIENDE'] == ambito_filtro) &(df_data_agende_con_ambito['Codice Agenda'].isin(unique_values)) ]
    df_data_agende_filtered = df_data_agende_filtered[['Codice Agenda','Descrizione Agenda','Grado di aderenza al grafo*','Grado di urgenza**','Sentinella']]

    four_tables_dict = four_tables.to_dict('records') if four_tables is not None else None
    df_prima_disponibilita_dict = df_prima_disponibilita_df.to_dict('records') if df_prima_disponibilita_df is not None else None
    df_data_agende_dict = df_data_agende_filtered.to_dict('records') if df_data_agende_filtered is not None else None
    df_daily_requests_in_minutes_per_agend_dict = df_daily_requests_in_minutes_per_agend.to_dict('records') 


        #return four_tables,direttore_reclustering_tempi_attesa,direttore_ore_ambulatorio_df,df_prima_disponibilita_2,df_data_agende,'/agende'
    return four_tables_dict,None,None,df_daily_requests_in_minutes_per_agend_dict,df_prima_disponibilita_dict,df_data_agende_dict,'/agende'   

@app.callback(
    Output('selected-ambiti-store', 'data'),
    Input('initial-dropdown', 'value')
)
def update_selected_code(selected_values):
    if selected_values:
        return selected_values
    else:
        return []
    
    
### Visualizzza l 'area territoriale scelta nella pagina del direttore non è un filtro    
@app.callback(
    Output('dropdown-menu', 'value'),
    Input('selected-ambiti-store', 'data')
)
def update_textarea(selected_value):
    if selected_value:
        print(selected_value)
        dizionario_decodifica = {"ASL Caserta A01" : "Caserta 01",
                                 "ASL Caserta A02" : "Caserta 02",
                                 "ASL Caserta A03" : "Caserta 03",
                                 "ASL Caserta A04" : "Caserta 04"}
        return dizionario_decodifica[selected_value]
    else:
        return ''
   
    
#@app.callback(
#    Output('dropdown-menu', 'options'),
#    Input('selected-ambiti-store', 'data')
#)
#def update_dropdown(selected_values):
#    if selected_values:
#        return [{'label': value_to_label[item['value']], 'value': item['value']} for item in selected_values]
#    else:
#        return [
#            {'label': 'Provincia di Caserta', 'value': 'ASL Caserta A04'},
#            {'label': 'Provincia di Caserta 1', 'value': 'ASL Caserta 1'},
#            {'label': 'Provincia di Caserta 2', 'value': 'ASL Caserta 2'},
#            {'label': 'Provincia di Caserta 3', 'value': 'ASL Caserta 3'}
#        ]


### LEGGE I VALORI UNICI ELLA COLONNA AMBITO E SI CREA LO STORE DELLE OPZIONI POSSIBILI
## 
def parse_contents(contents, filename):
    content_type, content_string = contents.split(',')

    decoded = base64.b64decode(content_string)
    try:
        if 'xls' in filename:
            # Usare pd.read_excel per leggere file excel
            df = pd.read_excel(io.BytesIO(decoded))
        elif 'csv' in filename:
            # Usare pd.read_csv per leggere file csv
            df = pd.read_csv(io.StringIO(decoded.decode('utf-8')))
        else:
            return None
    except Exception as e:
        print(e)
        return None

    return df

### DROPDOWN DELLA PRIMA PAGINA HOME PUNTO 3)
@app.callback(
    [Output('initial-dropdown', 'options'),
     Output('output-data-upload', 'children')],
    [Input('upload-data', 'contents')],
    [State('upload-data', 'filename'),
     State('initial-options', 'data')]
)
def update_output(contents, filename, stored_options):
    if contents is not None:
        df = parse_contents(contents, filename)
        # Effettua il join mantenendo solo le colonne desiderate
        
        df_joined = pd.merge(df, merged_df, 
                             on=['CODICE_AGENDA', 'CODICE_EROGANTE', 'STS11'], 
                             how='left')


        print("------------CHECK PRINT--------------------")
        print(df_joined)
        if df_joined is not None and 'AMBITI_AZIENDE' in df_joined.columns:
            unique_values = df_joined['AMBITI_AZIENDE'].unique()
            options = [{'label': val, 'value': val} for val in unique_values]
            dizionario_decodifica = {"ASL Caserta A01" : "Caserta 01",
                                 "ASL Caserta A02" : "Caserta 02",
                                 "ASL Caserta A03" : "Caserta 03",
                                 "ASL Caserta A04" : "Caserta 04"}
            options_provincia = [{'label': dizionario_decodifica[val], 'value': val} for val in unique_values]
            options_provincia_sorted = sorted(options_provincia, key=lambda x: x['label'])

            return options_provincia_sorted, ''
        else:
            return [], 'File in formato errato.'
    elif stored_options is not None:
        return stored_options, ''
    return [], ''

### SI RICORDA L AMBITO CHE HAI SALVATO 
@app.callback(
    Output('initial-options', 'data'),
    Input('initial-dropdown', 'options')
)
def remember_options(options):
    return options


### QUANDO TORNA INDIETRO SI RICORDA CHE IL FILE E'STATO CARICATO
@app.callback(
    Output('name', 'data'),
    Input('upload-container', 'children')
)
def remember_options(value):
    return value


### FA USCIRE LA SCRITTA NOME DEL FILE QUANDO FAI L UPLOAD 
@app.callback(
    Output('upload-container', 'children'),
    [Input('upload-data', 'contents')],
    [State('upload-data', 'filename'),
     State('name', 'data')]
)
def update_output(contents, filename, saved_filename):
    if contents is not None:
        return html.Div([
                    html.Br(),
                    html.Img(src='/assets/load_file.png', style={'height': '30px'}),
                    html.Div(filename, style={'textAlign': 'center', 'font-size': '14px', 'marginTop': '5px','margin-left':'10px','font-weight':'bold','textDecoration': 'underline'})
                ], style={'display': 'flex','alignItems': 'center','height':'50px'})
    #elif saved_filename is not None:
    #    return 
    #saved_filename
    return html.Div([
                html.Img(src='/assets/load_file.png', style={'height': '30px'}),
                html.Div('Trascina o seleziona il file', id='upload-text', style={'textAlign': 'center', 'font-size': '14px'})
            ], style={'display': 'flex', 'flexDirection': 'column', 'alignItems': 'center'}) 





### tabella data_agende
@app.callback(
    Output('table-new', 'rowData'),
    Output('table-new', 'columnDefs'),
    [ Input('store-df6', 'data')]
)
def update_table2(df_data_agende_dict):
    data_agende = pd.DataFrame.from_dict(df_data_agende_dict)
    #table_height = calculate_table_height(data_agende)
    print("----------CHECK---------")
    print(data_agende)
    data_agende["id"] = data_agende.index
    
    """def discrete_background_color_bins(df,columns):
        styleConditions = []

        for column in columns:
            if column == 'Sentinella':
                styleConditions.append({
                    "condition": "params.data['Sentinella'] == 'Si'",
                    "style": {"backgroundColor": "#FFFFA7",'display': 'flex',
            'justifyContent': 'center',
            'alignItems': 'center'},
                    
                })
                styleConditions.append({
                    "condition": "params.data['Sentinella'] == 'No'",
                    "style": {"backgroundColor": "#F0F6FF", "color": "black",'display': 'flex',
            'justifyContent': 'center',
            'alignItems': 'center'},
                })
            

        return styleConditions"""
    columns = [
        {'headerName': 'Codice Agenda', 'field': 'Codice Agenda'},
        {'headerName': 'Descrizione Agenda', 'field': 'Descrizione Agenda'},
        {'headerName': 'Sentinella', 'field': 'Sentinella'},
        {'headerName': 'Grado di aderenza al grafo*', 'field': 'Grado di aderenza al grafo*',"valueFormatter": { "function": "params.value ? Math.round((params.value * 100)) + '%' : '-'"}, "cellStyle": {
    "styleConditions": [
        {
            
            "condition": "params.value >= 0.001 && params.value <= 0.25",
            "style": {"color": "#FF0000","font-weight": "bold"}
        },
        {
            "condition": "params.value > 0.25",
            "style": {"color": "#00B050","font-weight": "bold"}
        }
    ],
    "defaultStyle": {"color": "#ece9e8"} 
}
},
        {'headerName': 'Grado di urgenza**', 'field': 'Grado di urgenza**',"valueFormatter": { "function": "params.value ?  Math.round((params.value * 100)) + '%' : '-'"}, "cellStyle": {
    "styleConditions": [
        {
            
            "condition": "params.value >= 0.001 && params.value <= 0.25",
            "style": {"color": "#00B050","font-weight": "bold"}
        },
        {
            "condition": "params.value > 0.25",
            "style": {"color": "#FF0000","font-weight": "bold"}
        },
        
    ],
    "defaultStyle": {"color": "#ece9e8"}  
}
,
        },
        {'headerName': 'Selezione agenda/e da prioritizzare',
        'field': 'row_selector',
        'checkboxSelection': True,
        "cellStyle": {'display': 'flex',
            'justifyContent': 'center',
            'alignItems': 'center',
            'backgroundColor':'#ECF2FE'}
            },
        

    ]
    # Creazione della tabella con Dash AG Grid
    return df_data_agende_dict,columns
    
# Callback per aprire e chiudere la modale
@app.callback(
    Output("modal-aderenza", "is_open"),
    [Input("image-click-aderenza", "n_clicks"), Input("close-aderenza", "n_clicks")],
    [State("modal-aderenza", "is_open")],
)
def toggle_modal(n1, n2, is_open):
    if n1 or n2:
        return not is_open
    return is_open
@app.callback(
    Output("modal-congruenza", "is_open"),
    [Input("image-click-congruenza", "n_clicks"), Input("close-congruenza", "n_clicks")],
    [State("modal-congruenza", "is_open")],
)
def toggle_modal(n1, n2, is_open):
    if n1 or n2:
        return not is_open
    return is_open
@app.callback(
    Output("modal-agende", "is_open"),
    [Input("image-click-agende", "n_clicks"), Input("close-agende", "n_clicks")],
    [State("modal-agende", "is_open")],
)
def toggle_modal(n1, n2, is_open):
    if n1 or n2:
        return not is_open
    return is_open


app.config.suppress_callback_exceptions=True
if __name__ == '__main__':
    app.run_server(debug=False)

